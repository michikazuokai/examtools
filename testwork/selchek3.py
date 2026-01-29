from openpyxl import load_workbook, Workbook
import re
import random
import unicodedata
from pathlib import Path
from datetime import datetime
import yaml

def normalize_code_cell(val):
    """空白や全角を正規化"""
    if val is None:
        return ""
    return unicodedata.normalize("NFKC", str(val)).strip()

def has_question_title(code: str) -> bool:
    """問題タイトル（# 【問題12-1】など）が含まれているか判定"""
    pattern = re.compile(r"#\s*【問題(\d+)(?:[-ー−–—](\d+))?】")
    for line in code.splitlines():
        normalized = normalize_code_cell(line)
        if pattern.search(normalized):
            return True
    return False

def get_cbn(num):
    if 1 <= num <= 26:
        return chr(ord('A') + num - 1)
    else:
        return None

def comtext(ptxt,sdict):
    clst=ptxt.split(",")
    dlst=[]
    for w in clst:
        w2=sdict[int(w)]
        dlst.append(str(w2))
#        dlst.append(get_cbn(w2))
    dlst.sort()
    return ','.join(dlst)

def write_list_to_excel_row(ws,col, frow, slst):
    """
    frow行目から、slstの値をA列に順番に書き込む
    """
    for idx, val in enumerate(slst):
        ws.cell(row=frow + idx, column=col, value=val)

def shuffle_once(lst):
    """
    渡されたリストをシャッフルし、新しいリストとして返す関数
    """
    new_list = lst[:]         # 元のリストをコピー
    random.shuffle(new_list) # シャッフル
    return new_list

def tokutenlst(ws,wsout,col):

    if not hasattr(tokutenlst, "check_answer"):
        tokutenlst.selflg = False  # answer/subanswer チェック用

    i = 1                      # openpyxl は 1 始まり
    n = ws.max_row
    selflg = False

    flg=True
    ansflg=False
    lst=[]
    frow=0
    qestcnt=0
    diffcnt=0
    while i <= n:
        code = normalize_code_cell(ws.cell(row=i, column=1).value)   # 1列目 (コード)
        if code =="e_exam":
            break

        if has_question_title(code):
#        if "# 【" in code.strip():
            tit=code
        if code in {"b_select", "b_subselect"}:
            tokutenlst.selflg = True
            selidx=0
        if code in {"e_select", "e_subselect"}:
            flg=True
            if tokutenlst.selflg:
#                print(frow,lst)
                slst=shuffle_once(lst)
                # A列に書き込み
                write_list_to_excel_row(wsout, col, frow, slst)
                sdic = dict(zip(lst, slst))
                #print(frow,sdic)
               
                lst=[]
                tokutenlst.selflg = False

        if code in {"select", "subselect"}:
            if tokutenlst.selflg:
                if flg:
                    #print(tit)
                    frow=i
                    flg=False
                lst.append(selidx+1)
                selidx+=1

        if code in {"b_answer", "b_subanswer"}:
            if ws.cell(row=i, column=2).value=="#select":
                ansflg=True
        if code in {"e_answer", "e_subanswer"}:
            ansflg=False
        if code in {"answer", "subanswer"}:
            if ansflg:
                c1=ws.cell(row=i, column=2).value
                try:
                    c1=int(c1)
                    c2=sdic[c1]
                except ValueError:
                    c2=comtext(c1,sdic)
                qestcnt+=1
                if c1!=c2:
                    diffcnt+=1
                #print(f"{qestcnt} {i} {c1} -> {c2}")
                wsout.cell(row=i, column=col, value=c2)

        i += 1
    p=diffcnt/qestcnt
    #print(f"col:{col} total:{qestcnt} diff:{diffcnt} percent:{p:.0%}")
    return p
# ===== 実行 =====
curdir = Path(__file__).parent.parent
excel_path = curdir / "input/試験問題.xlsx"
subject="2030402"
#subject="1010401"
wb = load_workbook(excel_path)
ws = wb[subject]

# データの作成
dt = datetime.now()
# シャッフル用のワークブック作成
wbout = load_workbook(curdir /"work/output_x.xlsx")
wsout = wbout[subject]

# --- 1. YAMLファイルの読み込み ---
yaml_path = curdir / "work/output_x.yaml"

with open(curdir /"work/output_x.yaml", "r", encoding="utf-8") as f:
    kekdic = yaml.safe_load(f)

# ファイルが存在し、中身があるか確認してから読み込む
try:
    with open(yaml_path, "r", encoding="utf-8") as f:
        # ファイルが空の場合にNoneが返されるのを防ぐため、空の辞書 {} を初期値とする
        kekdic = yaml.safe_load(f) 
        if kekdic is None:
            kekdic = {}
except FileNotFoundError:
    # ファイルが存在しない場合は、空の辞書として初期化する
    kekdic = {}
except Exception as e:
    print(f"YAMLファイルの読み込みエラー: {e}")
    # エラーが発生した場合は、処理を中断するか、空の辞書で続行するか判断
    kekdic = {}

# --- 2. 新しいデータの生成 ---
keklst=[]
for rcnt in range(3):
    diffp=0
    ecnt=0
    while diffp<0.95:
        diffp=tokutenlst(ws,wsout,rcnt+1)
        ecnt+=1
    s=get_cbn(rcnt+1)+":"+f"{diffp:.2%}"
    keklst.append(s)
#    print(f"col:{rcnt+1} percent:{diffp:.2%} 計算回数:{ecnt}")
dt = datetime.now()

# --- 3. 既存の辞書 kekdic に新しいデータを追加/更新する ---
# ⚠️ ここが最も重要な修正点です。
kekdic[subject] = {
    "exedate": dt.strftime('%Y-%m-%d %H:%M:%S'),
    "kekka": keklst
}

# --- 4. YAMLファイルへの書き出し ---
with open(yaml_path, "w", encoding="utf-8") as f:
    yaml.dump(kekdic, f, default_flow_style=False, allow_unicode=True)


# 必要に応じてシート名を変更
#wbout.save(curdir /"work/output_x.xlsx")