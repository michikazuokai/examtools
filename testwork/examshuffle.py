from openpyxl import load_workbook
from pathlib import Path
import random
import string


def shuffle(arr):
    indexed_arr = [(element, index) for index, element in enumerate(arr)]
    random.shuffle(indexed_arr)
    mapping = {}
    for new_idx, (element, orig_idx) in enumerate(indexed_arr):
        mapping[orig_idx] = new_idx
    return [v[0] for v in indexed_arr], mapping


alp2n = lambda c: ord(c) - ord('A')
n2alp = lambda n: chr(n + 65)


def shuffle_choices(ws):
    """
    ws: openpyxl worksheet
    """

    i = 1
    n = ws.max_row
    selst = []
    selrows = []      # select の row 番号を保持
    selflg = False
    svcode = ""
    nmap = {}

    while i <= n:
        code = ws.cell(row=i, column=1).value
        data = ws.cell(row=i, column=2).value

        if code is None:
            i += 1
            continue

        # --------------------------------------------------
        # 問題名コード
        # --------------------------------------------------
        if "【" in str(code):
            svcode = code

        # --------------------------------------------------
        # 大問 b_select
        # --------------------------------------------------
        if code == "b_select":
            print(svcode)
            print("--- select ---")
            selflg = True
            selst = []
            selrows = []

        # select行の収集
        if selflg and code == "select":
            selst.append(data)
            selrows.append(i)

        # e_select → シャッフルして書き換え
        if selflg and code == "e_select":
            print(selst)
            narr, nmap = shuffle(selst)
            print("shuffled:", narr)
            print("map:", nmap)

            # ★ Excel に書き戻す
            for idx, rowno in enumerate(selrows):
                ws.cell(row=rowno, column=2).value = narr[idx]

        # answer → 新ラベルを書き換え
        if selflg and code == "answer":
            al = str(data).split(',')
            new_labels = []
            for v in al:
                ov = alp2n(v)
                nv = nmap[ov]
                new_labels.append(n2alp(nv))

            # ★ Excel に書き戻す
            ws.cell(row=i, column=2).value = ",".join(new_labels)

        if selflg and code == "e_answer":
            print("@@@ answer end ---")
            selflg = False
            selst = []
            selrows = []
            nmap = {}

        # --------------------------------------------------
        # 小問 b_subselect
        # --------------------------------------------------
        if code == "b_subselect":
            print(svcode)
            print("--- sub select ---")
            selflg = True
            selst = []
            selrows = []

        if selflg and code == "subselect":
            selst.append(data)
            selrows.append(i)

        if selflg and code == "e_subselect":
            print(selst)
            narr, nmap = shuffle(selst)
            print("shuffled:", narr)
            print("map:", nmap)

            # ★ 小問 select の書き換え
            for idx, rowno in enumerate(selrows):
                ws.cell(row=rowno, column=2).value = narr[idx]

        if selflg and code == "subanswer":
            al = str(data).split(',')
            new_labels = []
            for v in al:
                ov = alp2n(v)
                nv = nmap[ov]
                new_labels.append(n2alp(nv))

            # ★ 小問 subanswer の書き換え
            ws.cell(row=i, column=2).value = ",".join(new_labels)

        if selflg and code == "e_subanswer":
            print("@@@ sub answer end ---")
            selflg = False
            selst = []
            selrows = []
            nmap = {}

        i += 1


# ===== 実行例 =====
curdir = Path(__file__).parent
excel_path = curdir / "exam_1.xlsx"

wb = load_workbook(excel_path)
ws = wb.active

shuffle_choices(ws)

excel_path2 = curdir / "exam_2.xlsx"
wb.save(excel_path2)
print("書き換え完了:", excel_path2)