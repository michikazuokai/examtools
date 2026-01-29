from openpyxl import load_workbook
import json
from pathlib import Path
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from openpyxl.styles import Font
import sqlite3
import sys

font12 = Font(
    name='ＭＳ ゴシック',    # フォント名
    size=12)

font20 = Font(
    name='ＭＳ ゴシック',    # フォント名
    size=18)

def get_nenji_by_subno(sub_no):
    """subNoからnenjiを取得"""
    db_path = "/Volumes/NBPlan/TTC/カルテ管理/2025/DB/classdb.db"
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT nennji FROM class WHERE subNo = ?", (sub_no,))
        result = cursor.fetchone()
        return str(result[0]) if result else None
    except sqlite3.Error as e:
        print(f"エラー: {e}")
        return None
    finally:
        conn.close()


def special_cumsum(arr):
    if len(arr) == 1:
        return [1]
    result = [1]

    for i in range(1, len(arr)):
        prev_val = arr[i-1]
        curr = arr[i]
        if curr == prev_val:
            # 同じ値 → その値を加算
            result.append(result[-1] + curr)
        else:
            # 値が変わったとき
            if prev_val == 1:
                result.append(result[-1] + 1)
            else:
                result.append(result[-1] + prev_val)
    return result


def keisen(ws, row,start_col, haba,lvl,ans):
    ws.cell(row=row, column=start_col).value = lvl
    ws.cell(row=row+1, column=start_col).value = ans
    ws.cell(row=row, column=start_col).font=font12
    ws.cell(row=row+1, column=start_col).font=font20
    if haba == 1:
        side = Side(style='thin', color='000000')
        border = Border(top=side, bottom=side, left=side, right=side)
        ws.cell(row=row, column=start_col).border = border
        ws.cell(row=row+1, column=start_col).border = border
        ws.cell(row=row, column=start_col).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row+1, column=start_col).alignment = Alignment(horizontal='center', vertical='center')
        
    else:
        for r in range(row, row+2):
            thin = Side(style='thin', color='000000')
            for col in range(start_col, start_col+haba ):
                cell = ws.cell(row=r, column=col)
                if col == start_col:
                    # 左端セル：上・下・左
                    cell.border = Border(top=thin, bottom=thin, left=thin)
                elif col == (start_col+haba-1 ):
                    # 右端セル：上・下・右
                    cell.border = Border(top=thin, bottom=thin, right=thin)
                else:
                    # 中間セル：上・下のみ
                    cell.border = Border(top=thin, bottom=thin)
            ws.merge_cells(start_row=r, start_column=start_col, end_row=r, end_column=start_col+haba-1)
            ws.cell(row=r, column=start_col).alignment = Alignment(wrap_text=True,horizontal='center', vertical='center')

#-------------------------------------------------------------------
if len(sys.argv) < 2:
    print("Usage: python anstest_ans.py <科目番号>")
    sys.exit(1)
subject = sys.argv[1]  #シート名：科目番号

curdir = Path(__file__).parent
excel_path = curdir / "anssheet_template.xlsx"
wb = load_workbook(excel_path)
ws = wb["template"]

#subject="2030402"
po=curdir.parent / "work" / f"answers_{subject}.json"
with open(po, "r", encoding="utf-8") as f:
    outjson = json.load(f)

# 試験タイトルのセット
ver=outjson["versions"][0]["version"]
title=outjson["versions"][0]["questions"][0]["title"]
#
ntitle=f"{title} ({ver})   履修判定試験"
newv=ws["A1"].value.replace("##title##",ntitle)
ws["A1"].value=newv
#
nenji=get_nenji_by_subno(subject)
newn=ws["A2"].value.replace("##nenji##",nenji)
ws["A2"].value=newn

ws["A1"].value=newv

for ir,al in enumerate(outjson["versions"][0]["questions"][1:]):
    a = al["width"]
    b=[int(v) for v in al["width"]]
    result=special_cumsum(b)
    rw=4+ir*3
    for i,v in enumerate(result):
        lvl=al["label"][i]
        ans=""
        keisen(ws, rw,v,int(a[i]),lvl,ans)
ws.title = "解答用紙"


# 3. シートをコピーする
ws_kaito = wb.copy_worksheet(ws)

# 4. 新しいシート名を設定
ws_kaito.title = "模範解答"
for ir,al in enumerate(outjson["versions"][0]["questions"][1:]):
    a = al["width"]
    b=[int(v) for v in al["width"]]
    result=special_cumsum(b)
    rw=4+ir*3
    for i,v in enumerate(result):
        lvl=al["label"][i]
        ans=al["answer"][i]
        keisen(ws_kaito, rw,v,int(a[i]),lvl,ans)
# 4. シート名を変更したい場合

title = outjson["versions"][0]["questions"][0]["title"]
ver=outjson["versions"][0]['version']
fname=f"{subject}_{title}_{ver}解答用紙.xlsx"
wb.save(curdir /fname)