import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

def create_excel_from_answers(json_path: str | Path, out_xlsx: str | Path) -> None:
    """
    answers_XXXX.json（versionmode=multi, versions[].questions[] 構造）
    から、バージョンごとに 1 シートずつ解答一覧を作成する。

    - questions[0] … ヘッダ情報（title 等）: Excel には出力しない
    - questions[1:] … 各要素が 1 行分の解答ブロック
      - block["label"]  … 上段に横展開
      - block["answer"] … 下段に横展開
      - block["width"]  … 列幅の目安（列方向に最大値を採用）
    """

    json_path = Path(json_path)
    out_xlsx = Path(out_xlsx)

    with json_path.open("r", encoding="utf-8") as f:
        data = json.load(f)

    if data.get("versionmode") != "multi":
        raise ValueError("この関数は versionmode == 'multi' の JSON を前提にしています。")

    wb = Workbook()
    # デフォルトシート削除
    wb.remove(wb.active)

    # ─────────────────────────────
    # 各バージョンごとにシートを作成
    # ─────────────────────────────
    for ventry in data["versions"]:
        ver_name = ventry.get("version", "A")
        qlist = ventry.get("questions", [])
        if not qlist:
            continue

        # questions[0] はヘッダ（title/width/height…）なのでスキップ
        header = qlist[0]
        blocks = qlist[1:]

        ws = wb.create_sheet(title=ver_name)

        # 列幅を決めるために最大 width を列ごとに記録しておく
        max_width_per_col = {}

        # block_index: 0,1,2,... → Excel では 1ブロックにつき 2行使う
        for block_idx, block in enumerate(blocks):
            labels  = block.get("label", [])
            answers = block.get("answer", [])
            widths  = block.get("width", [])

            # ラベル行と解答行
            label_row = block_idx * 2 + 1   # 1,3,5,...
            ans_row   = block_idx * 2 + 2   # 2,4,6,...

            # 各要素を横方向に展開
            for i, lbl in enumerate(labels):
                col = i + 1  # 1-based

                # 上段：ラベル
                c_label = ws.cell(row=label_row, column=col, value=lbl)
                c_label.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                # 下段：解答文字列
                ans_text = answers[i] if i < len(answers) else ""
                c_ans = ws.cell(row=ans_row, column=col, value=ans_text)
                # 解答は折り返しありにしておくと長文でも収まる
                c_ans.alignment = Alignment(wrap_text=True, vertical="top")

                # 列幅候補の更新（width があれば）
                if i < len(widths):
                    w = widths[i]
                    # 同じ列で最大値を保持
                    if col not in max_width_per_col or max_width_per_col[col] < w:
                        max_width_per_col[col] = w

            # 行の高さはざっくり固定しておく（必要なら調整）
            ws.row_dimensions[label_row].height = 18
            ws.row_dimensions[ans_row].height = 30

        # 列の幅を width 値から決定（倍率は見た目用に調整）
        for col_idx, base_w in max_width_per_col.items():
            col_letter = get_column_letter(col_idx)
            # base_w は 1.0, 3.0, 5.0 などなので、少し大きめに
            ws.column_dimensions[col_letter].width = base_w * 3.5

    # 保存
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)
    print(f"✔ Excel 出力完了: {out_xlsx}")

curdir = Path(__file__).parent.parent
json_path = curdir / "work/answers_2022001.json"
excel_path = curdir / "work/ans_2022001.xlsx"

# 実行例
create_excel_from_answers(json_path,excel_path)
#create_excel_from_answers(json_path,excel_path)


