from openpyxl import load_workbook
import re
import unicodedata
from pathlib import Path



curdir = Path(__file__).parent.parent
excel_path = curdir / "input/試験問題.xlsx"
wb = load_workbook(excel_path)
ws = wb["1010401"]


# パターン
pattern = re.compile(r"#\s*【問題(\d+)[\-ー−–—]?\d*】")
#pattern = re.compile(r"#\s*【問題\d+】")

# マッチチェック
matches = []
non_matches = []
for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
    for cell in row:
        if isinstance(cell, str):
            normalized = unicodedata.normalize("NFKC", cell).strip()
            if pattern.search(normalized):
                matches.append((i, normalized))
            else:
                if "【問題" in normalized:
                    non_matches.append((i, normalized))

# 結果表示
import pandas as pd
df_non_matches = pd.DataFrame(non_matches, columns=["行番号", "マッチしなかった文字列"])
print(df_non_matches)