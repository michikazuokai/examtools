import argparse
import openpyxl
import re
import sys
from pathlib import Path

# ============================================================
# validate_excelv2.py
#   - format=v2 専用バリデータ
#   - 既存の validate_excel.py は変更しない
#
# v2 ルール（本スレッド確定事項のうち validate が担担保する範囲）
#   1) b_exam の C列に "format=v2" があること
#   2) b_question の C〜H列（QID/orderB/PB/LS）の型・必須・重複
#   3) LINESPACE 行タグは b_subgroup 内だけ許可（A案）
#   4) PAGEBREAK 行タグは v2 では禁止（列PB_*_afterへ移行する前提）
#   5) b_answer / b_subanswer の C列は difficulty（LOW/MID/HIGH）
#   6) answer / subanswer の C列は配点（整数）
#   7) #select の場合も answer/subanswer の配点チェックは行う
# ============================================================

# ----------------------------
# ヘルパー
# ----------------------------
def is_integer(v) -> bool:
    try:
        if v is None:
            return False
        int(str(v).strip())
        return True
    except Exception:
        return False

def is_number(v) -> bool:
    try:
        if v is None or str(v).strip() == "":
            return False
        float(str(v).strip())
        return True
    except Exception:
        return False

def is_2digits(v) -> bool:
    """
    "#select" の answer/subanswer の形式：1〜2桁数字、またはカンマ区切り
    例: "3" / "12" / "1,3" / "2,10"
    """
    if v is None:
        return False
    s = str(v).strip()
    if s == "":
        return False
    # 1〜2桁数字のカンマ区切り
#    return re.fullmatch(r"\d{1,2}(,\d{1,2})*", s) is not None
    return re.fullmatch(r"\d+(,\d+)*", s) is not None

def is_flag(v) -> bool:
    """PB_*_after 用：空 or 1/true/y/yes 等"""
    if v is None:
        return True
    s = str(v).strip().lower()
    if s == "":
        return True
    return s in {"1", "true", "t", "y", "yes"}

def is_difficulty(v) -> bool:
    if v is None:
        return False
    s = str(v).strip().upper()
    return s in {"LOW", "MID", "HIGH"}

def normalize_tag(v) -> str:
    if v is None:
        return ""
    return str(v).strip()

def cell(row, col_index_1based):
    """openpyxlの values_only 行（tuple）から 1-based で安全に取得"""
    i = col_index_1based - 1
    if i < 0:
        return None
    if i >= len(row):
        return None
    return row[i]

# ----------------------------
# 構造ルール（v1と同等）
# ※ LINESPACE/PAGEBREAK は v2で別処理（構造チェックから除外）
# ----------------------------
ALLOWED_CHILDREN = {
    "b_exam": {"examtitle", "b_examnote", "subject", "fsyear", "ansnote", "anssize",
               "b_question", "qpattern"},
    "b_examnote": {"examnote"},
    "b_question": {"question", "image", "sline", "b_multiline",
                   "b_select", "b_code", "b_subgroup",
                   "b_answer"},
    "b_multiline": {"text"},
    "b_select": {"select"},
    "b_subselect": {"subselect"},
    "b_code": {"code"},
    "b_subcode": {"subcode"},
    "b_subgroup": {"b_subquest"},
    "b_subquest": {"subquest", "subimage", "subsline", "b_submultiline",
               "b_subselect", "b_subcode", "b_subanswer"},
    "b_submultiline": {"subtext"},
    "b_answer": {"answer"},
    "b_subanswer": {"subanswer"},
}

CLOSING_TAGS = {
    "e_examnote": "b_examnote",
    "e_question": "b_question",
    "e_multiline": "b_multiline",
    "b_submultiline": "e_submultiline",
    "e_select": "b_select",
    "e_subselect": "b_subselect",
    "e_code": "b_code",
    "e_subgroup": "b_subgroup",
    "e_subquest": "b_subquest",
    "e_answer": "b_answer",
    "e_subanswer": "b_subanswer",
}

REQUIRED_CHILDREN = {
    "b_exam": {"examtitle", "subject", "fsyear"},
    "b_question": {"question"},
    "b_select": {"select"},
    "b_subselect": {"subselect"},
    "b_answer": {"answer"},
    "b_subanswer": {"subanswer"},
}

def is_comment(tag: str) -> bool:
    return tag.startswith("#")

def check_structure(tag, stack, rownum, errors):
    """
    既存validateと同様：b_〜e_の整合＋子タグ許可
    """
    if not hasattr(check_structure, "block_children"):
        check_structure.block_children = {}

    if tag.startswith("b_"):
        stack.append(tag)
        if tag not in check_structure.block_children:
            check_structure.block_children[tag] = set()
        return

    if tag.startswith("e_"):
        if not stack:
            errors.append(f"Row {rownum}: '{tag}' without matching opening tag")
            return
        expected = "b_" + tag[2:]
        if stack[-1] != expected:
            errors.append(f"Row {rownum}: '{tag}' does not match open tag '{stack[-1]}'")
            return
        open_tag = stack.pop()

        # 必須子タグチェック
        required = REQUIRED_CHILDREN.get(open_tag)
        if required:
            seen = check_structure.block_children.get(open_tag, set())
            missing = required - seen
            if missing:
                errors.append(
                    f"Row {rownum}: '{open_tag}' missing required child tags: {sorted(missing)}"
                )
        return

    # 中身タグ
    if stack:
        parent = stack[-1]
        allowed = ALLOWED_CHILDREN.get(parent, set())
        if tag not in allowed:
            errors.append(f"Row {rownum}: Invalid child tag '{tag}' inside '{parent}'")
        else:
            check_structure.block_children.setdefault(parent, set()).add(tag)
    else:
        # ルートで許されるタグ
        if tag not in {"b_exam"}:
            errors.append(f"Row {rownum}: Tag '{tag}' must appear inside a block")

# ----------------------------
# 値チェック（v2）
# ----------------------------
def check_values_v2(tag, row, rownum, stack, errors, state):
    """
    state:
      - format_ok: b_examのC列でformat=v2を確認したか
      - check_answer: 直前b_answer/b_subanswerが #select だったか
      - seen_qids: set
      - seen_orderB: set
    """
    # 1) v2の開始判定
    if tag == "b_exam":
        fmt = cell(row, 3)
        if str(fmt).strip().lower() != "format=v2":
            errors.append(
                f"Row {rownum}: b_exam C列 must be 'format=v2' (got '{fmt}')"
            )
        else:
            state["format_ok"] = True
        return

    # 2) v2では PAGEBREAK 行タグは禁止
    if tag == "PAGEBREAK":
        errors.append(f"Row {rownum}: PAGEBREAK is not allowed in format=v2 (use PB_*_after columns on b_question).")
        return

    # 3) v2では LINESPACE 行タグは b_subgroup 内だけ許可（A案）
    if tag == "LINESPACE":
        if "b_subgroup" not in stack:
            errors.append(f"Row {rownum}: LINESPACE is allowed only inside b_subgroup in format=v2.")
        num = cell(row, 2)
        if num is None or str(num).strip() == "" or not is_number(num):
            errors.append(f"Row {rownum}: LINESPACE B列 must be a number (got '{num}').")
        return

    # 4) b_question の v2列（C〜H）
    if tag == "b_question":
        qid = cell(row, 3)
        orderB = cell(row, 4)
        pb_a = cell(row, 5)
        ls_a = cell(row, 6)
        pb_b = cell(row, 7)
        ls_b = cell(row, 8)

        # QID
        if qid is None or str(qid).strip() == "":
            errors.append(f"Row {rownum}: b_question C列(QID) is required.")
        else:
            qid_s = str(qid).strip()
            if qid_s in state["seen_qids"]:
                errors.append(f"Row {rownum}: duplicated QID '{qid_s}'.")
            state["seen_qids"].add(qid_s)

        # orderB
        if orderB is None or str(orderB).strip() == "":
            errors.append(f"Row {rownum}: b_question D列(orderB) is required.")
        elif not is_integer(orderB):
            errors.append(f"Row {rownum}: b_question D列(orderB) must be integer (got '{orderB}').")
        else:
            ob = int(str(orderB).strip())
            if ob in state["seen_orderB"]:
                errors.append(f"Row {rownum}: duplicated orderB '{ob}'.")
            state["seen_orderB"].add(ob)

        # PB/LS
        if not is_flag(pb_a):
            errors.append(f"Row {rownum}: PB_A_after(E列) must be empty or a flag (got '{pb_a}').")
        if ls_a is not None and str(ls_a).strip() != "" and not is_number(ls_a):
            errors.append(f"Row {rownum}: LS_A_after(F列) must be a number (got '{ls_a}').")
        if not is_flag(pb_b):
            errors.append(f"Row {rownum}: PB_B_after(G列) must be empty or a flag (got '{pb_b}').")
        if ls_b is not None and str(ls_b).strip() != "" and not is_number(ls_b):
            errors.append(f"Row {rownum}: LS_B_after(H列) must be a number (got '{ls_b}').")
        return

    # 5) b_answer / b_subanswer：#select判定＋difficulty(C列)
    if tag in {"b_answer", "b_subanswer"}:
        bval = cell(row, 2)
        state["check_answer"] = (str(bval).strip().lower() == "#select")
        diff = cell(row, 3)
        if diff is None or str(diff).strip() == "":
            errors.append(f"Row {rownum}: {tag} C列(difficulty) is required (LOW/MID/HIGH).")
        elif not is_difficulty(diff):
            errors.append(f"Row {rownum}: {tag} C列(difficulty) must be LOW/MID/HIGH (got '{diff}').")
        return

    # 6) answer / subanswer：#select時の形式＋点数（C列）
    if tag in {"answer", "subanswer"}:
        ans = cell(row, 2)
        point = cell(row, 3)

        if state.get("check_answer", False):
            if not is_2digits(ans):
                errors.append(
                    f"Row {rownum}: Invalid answer value '{ans}' for tag '{tag}' "
                    f"(#select mode: must be 1–2 digits or comma-separated 1–2 digits)"
                )

        # 点数（C列）は常に検証
        if point is None or str(point).strip() == "":
            errors.append(f"Row {rownum}: {tag} C列(point) is required (integer).")
        elif not is_integer(point):
            errors.append(f"Row {rownum}: {tag} C列(point) must be integer (got '{point}').")

        state["check_answer"] = False
        return

    # 7) その他の既存ルール（最低限）
    if tag == "examtitle":
        if cell(row, 2) is None or str(cell(row, 2)).strip() == "":
            errors.append(f"Row {rownum}: examtitle B列 is required.")
    if tag == "subject":
        if cell(row, 2) is None or str(cell(row, 2)).strip() == "":
            errors.append(f"Row {rownum}: subject B列 is required.")
    if tag == "fsyear":
        if cell(row, 2) is None or str(cell(row, 2)).strip() == "":
            errors.append(f"Row {rownum}: fsyear B列 is required.")

# ----------------------------
# 得点一覧（簡易）
# ----------------------------
def tokutenlst(ws):
    out = []
    total = 0
    cur = None
    cur_sum = 0
    in_q = False

    for i in range(1, ws.max_row + 1):
        code = ws.cell(row=i, column=1).value
        if code is None:
            continue
        code = str(code).strip()

        if code.startswith("#") and "【問題" in code:
            cur = code
            cur_sum = 0
            in_q = True
            continue

        if in_q and code in {"answer", "subanswer"}:
            ten = ws.cell(row=i, column=3).value
            if ten is not None and str(ten).strip() != "" and is_integer(ten):
                cur_sum += int(str(ten).strip())
                out.append(f"{cur}  {int(str(ten).strip())}")
###                out.append(f"{cur}  {cur_sum}")
                total += int(str(ten).strip())
###                total += cur_sum

        if in_q and code == "e_question":
            # out.append(f"{cur}  {cur_sum}")
            # total += cur_sum
            in_q = False

    out.append(f"合計  {total}")
    return out

def validate_excelv2(path: str, shname="Sheet1"):
    wb = openpyxl.load_workbook(path)
    if shname not in wb.sheetnames:
        raise ValueError(f"Sheet '{shname}' not found. Available: {wb.sheetnames}")
    ws = wb[shname]

    errors = []
    stack = []

    state = {
        "format_ok": False,
        "check_answer": False,
        "seen_qids": set(),
        "seen_orderB": set(),
    }

    if hasattr(check_structure, "block_children"):
        delattr(check_structure, "block_children")

    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        tag = normalize_tag(cell(row, 1))
        if tag == "" or is_comment(tag):
            continue

        if tag in {"LINESPACE", "PAGEBREAK"}:
            check_values_v2(tag, row, idx, stack, errors, state)
            continue

        check_structure(tag, stack, idx, errors)
        check_values_v2(tag, row, idx, stack, errors, state)

    if stack:
        errors.append(f"EOF: Unclosed tags remain on stack: {stack}")

    return errors, tokutenlst(ws)

from pathlib import Path
import argparse
import sys

def main():
    # 1. パスの基準設定
    # スクリプトの2階層上を curdir とし、その下の input フォルダを既定とする
    curdir = Path(__file__).resolve().parent.parent
    input_dir = curdir / "input"

    # 2. 引数の定義
    parser = argparse.ArgumentParser(description="試験問題バリデーションツール")

    # シート名：位置引数（Positional Argument）として必須にする
    parser.add_argument("sheet", help="シート名 (例: 1020201)")

    # --path：オプション引数（既定値：試験問題.xlsx）
    parser.add_argument("--path", default="試験問題.xlsx", help="エクセルファイル名")

    args = parser.parse_args()

    # 3. エクセルファイルのフルパスを構築
    # --path でファイル名だけが指定された場合も、input ディレクトリと結合する
    excel_path = input_dir / args.path

    # デバッグ情報の表示（必要に応じて）
    print(f"Target Path: {excel_path}")
    print(f"Target Sheet: {args.sheet}")

    # 4. バリデーションの実行
    # excel_path は Path オブジェクトなので、文字列に変換して関数に渡す
    errs, toklst = validate_excelv2(str(excel_path), args.sheet)

    if errs:
        print("Validation errors:")
        for e in errs:
            print(" -", e)
        sys.exit(1)
    else:
        print(f"Validation OK! (format=v2, sheet={args.sheet})")
        
        # トークンリストの表示（必要に応じて）
        for v in toklst:
            print(v)
        sys.exit(0)

if __name__ == "__main__":
    main()