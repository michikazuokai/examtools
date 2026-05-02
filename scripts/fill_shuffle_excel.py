# scripts/fill_shuffle_excel.py
from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from utils import get_exam_path


# B案：
# pattern[i] は「元の i+1 番の選択肢が、新しい何番へ移動するか」
# 例: [2,3,4,1]
# 元1 -> 新2, 元2 -> 新3, 元3 -> 新4, 元4 -> 新1
PATTERNS = [
    [2, 3, 4, 1],
    [3, 4, 1, 2],
]

# 何回も実行できるように、最初にG列をクリアする対象タグ
# b_question のG列は PB_B_after の可能性があるため消さない
TARGET_CLEAR_TAGS = {"select", "subselect", "answer", "subanswer"}


def norm_tag(value: Any) -> str:
    """A列タグを文字列化して整える。"""
    return "" if value is None else str(value).strip()


def parse_answer_numbers(value: Any) -> list[int]:
    """
    answer / subanswer のB列にある元の正解番号を list[int] にする。

    対応例:
      2       -> [2]
      "2"     -> [2]
      "1,3"   -> [1, 3]
      "1、3"  -> [1, 3]
      "1 3"   -> [1, 3]

    数値以外が混じる場合は変換対象外として [] を返す。
    """
    if value is None:
        return []

    text = str(value).strip()
    if not text:
        return []

    # 全角数字を半角へ
    trans = str.maketrans("１２３４５６７８９０", "1234567890")
    text = text.translate(trans)

    # 区切りをカンマに統一
    text = text.replace("，", ",").replace("、", ",")
    text = re.sub(r"\s+", ",", text)

    parts = [p.strip() for p in text.split(",") if p.strip()]
    if not parts:
        return []

    nums: list[int] = []
    for p in parts:
        if not re.fullmatch(r"\d+", p):
            return []

        n = int(p)

        # 今回は4択固定
        if n < 1 or n > 4:
            return []

        nums.append(n)

    return nums


def convert_answers(old_answers: list[int], pattern: list[int]) -> list[int]:
    """
    B案の正解番号変換。

    pattern:
      元の選択肢番号ごとの移動先。
      例: [2,3,4,1]

    old_answers:
      元の正解番号。
      例: [2] または [1,3]
    """
    return sorted(pattern[a - 1] for a in old_answers)


def format_answer_numbers(values: list[int]) -> Any:
    """
    G列に書く値を作る。
    単一正解なら数値、複数回答なら '1,3' の文字列。
    """
    values = sorted(values)

    if len(values) == 1:
        return values[0]

    return ",".join(str(v) for v in values)


def clear_g_column(ws) -> int:
    """
    select / subselect / answer / subanswer のG列だけをクリアする。

    b_question のG列は PB_B_after、
    b_subquest などのG列も別用途の可能性があるため消さない。
    """
    count = 0

    for row_no in range(1, ws.max_row + 1):
        tag = norm_tag(ws.cell(row_no, 1).value)

        if tag in TARGET_CLEAR_TAGS:
            if ws.cell(row_no, 7).value is not None:
                count += 1
            ws.cell(row_no, 7).value = None

    return count


def apply_pattern_to_choice_rows(ws, rows: list[int], pattern: list[int]) -> None:
    """
    select / subselect の4行にG列の変換番号をセットする。
    """
    if len(rows) != 4:
        raise ValueError(f"選択肢が4行ではありません: rows={rows}")

    for row_no, value in zip(rows, pattern):
        ws.cell(row_no, 7).value = value


def fill_shuffle_for_sheet(ws) -> dict[str, int]:
    """
    b_question〜e_question、b_subquest〜e_subquest の範囲を意識して、
    G列にB版用シャッフル番号と変換後正解番号をセットする。

    select:
      b_question〜e_question 内の選択肢として扱う

    answer:
      同じ b_question〜e_question 内の select のパターンで変換する

    subselect:
      b_subquest〜e_subquest 内の選択肢として扱う

    subanswer:
      同じ b_subquest〜e_subquest 内の subselect のパターンで変換する
    """
    stats = {
        "question_choice_groups": 0,
        "subquestion_choice_groups": 0,
        "choice_rows": 0,
        "answer_rows": 0,
        "subanswer_rows": 0,
        "warnings": 0,
    }

    # 大問・小問の状態
    in_question = False
    in_subquest = False

    # answer / subanswer ブロックの状態
    in_answer_block = False
    in_subanswer_block = False

    # b_answer / b_subanswer のB列が #select のときだけ変換対象
    answer_is_select = False
    subanswer_is_select = False

    # 現在の大問・小問で使った変換パターン
    current_question_pattern: list[int] | None = None
    current_subquest_pattern: list[int] | None = None

    # select / subselect の行を一時保持
    select_rows: list[int] = []
    subselect_rows: list[int] = []

    # パターンを交互に使うためのカウンタ
    # select と subselect を別カウントにする
    question_select_group_index = 0
    subquestion_select_group_index = 0

    def flush_select_rows(current_row: int) -> None:
        """
        大問内 select の4行を処理する。
        """
        nonlocal select_rows
        nonlocal current_question_pattern
        nonlocal question_select_group_index

        if not select_rows:
            return

        if len(select_rows) != 4:
            print(f"警告: select が4行ではありません: rows={select_rows}, near row={current_row}")
            stats["warnings"] += 1
            select_rows = []
            current_question_pattern = None
            return

        pattern = PATTERNS[question_select_group_index % len(PATTERNS)]
        apply_pattern_to_choice_rows(ws, select_rows, pattern)

        current_question_pattern = pattern
        question_select_group_index += 1

        stats["question_choice_groups"] += 1
        stats["choice_rows"] += 4

        select_rows = []

    def flush_subselect_rows(current_row: int) -> None:
        """
        小問内 subselect の4行を処理する。
        """
        nonlocal subselect_rows
        nonlocal current_subquest_pattern
        nonlocal subquestion_select_group_index

        if not subselect_rows:
            return

        if len(subselect_rows) != 4:
            print(f"警告: subselect が4行ではありません: rows={subselect_rows}, near row={current_row}")
            stats["warnings"] += 1
            subselect_rows = []
            current_subquest_pattern = None
            return

        pattern = PATTERNS[subquestion_select_group_index % len(PATTERNS)]
        apply_pattern_to_choice_rows(ws, subselect_rows, pattern)

        current_subquest_pattern = pattern
        subquestion_select_group_index += 1

        stats["subquestion_choice_groups"] += 1
        stats["choice_rows"] += 4

        subselect_rows = []

    for row_no in range(1, ws.max_row + 1):
        tag = norm_tag(ws.cell(row_no, 1).value)

        # ----------------------------
        # 大問開始・終了
        # ----------------------------
        if tag == "b_question":
            # 念のため前の未処理を閉じる
            flush_select_rows(row_no)
            flush_subselect_rows(row_no)

            in_question = True
            current_question_pattern = None
            select_rows = []

            # 大問ごとに select パターンをリセットしたい場合はここで0にする
            # 今回はシート全体で交互にするのでリセットしない
            continue

        if tag == "e_question":
            flush_select_rows(row_no)
            flush_subselect_rows(row_no)

            in_question = False
            in_answer_block = False
            answer_is_select = False
            current_question_pattern = None
            select_rows = []
            continue

        # ----------------------------
        # 小問開始・終了
        # ----------------------------
        if tag == "b_subquest":
            flush_select_rows(row_no)
            flush_subselect_rows(row_no)

            in_subquest = True
            current_subquest_pattern = None
            subselect_rows = []
            continue

        if tag == "e_subquest":
            flush_subselect_rows(row_no)

            in_subquest = False
            in_subanswer_block = False
            subanswer_is_select = False
            current_subquest_pattern = None
            subselect_rows = []
            continue

        # ----------------------------
        # select
        # ----------------------------
        if tag == "select":
            if not in_question:
                print(f"警告: row {row_no} の select が b_question〜e_question の外にあります")
                stats["warnings"] += 1
                continue

            select_rows.append(row_no)

            if len(select_rows) == 4:
                flush_select_rows(row_no)

            continue

        # select 以外が来たら、大問selectの未処理を閉じる
        # ただし、selectが4行単位でない場合の検出にもなる
        if select_rows and tag != "select":
            flush_select_rows(row_no)

        # ----------------------------
        # subselect
        # ----------------------------
        if tag == "subselect":
            if not in_subquest:
                print(f"警告: row {row_no} の subselect が b_subquest〜e_subquest の外にあります")
                stats["warnings"] += 1
                continue

            subselect_rows.append(row_no)

            if len(subselect_rows) == 4:
                flush_subselect_rows(row_no)

            continue

        # subselect 以外が来たら、小問subselectの未処理を閉じる
        if subselect_rows and tag != "subselect":
            flush_subselect_rows(row_no)

        # ----------------------------
        # b_answer / e_answer
        # ----------------------------
        if tag == "b_answer":
            in_answer_block = True
            marker = ws.cell(row_no, 2).value
            answer_is_select = str(marker).strip() == "#select" if marker is not None else False
            continue

        if tag == "e_answer":
            in_answer_block = False
            answer_is_select = False
            continue

        # ----------------------------
        # b_subanswer / e_subanswer
        # ----------------------------
        if tag == "b_subanswer":
            in_subanswer_block = True
            marker = ws.cell(row_no, 2).value
            subanswer_is_select = str(marker).strip() == "#select" if marker is not None else False
            continue

        if tag == "e_subanswer":
            in_subanswer_block = False
            subanswer_is_select = False
            continue

        # ----------------------------
        # answer
        # ----------------------------
        if tag == "answer":
            if not in_answer_block or not answer_is_select:
                continue

            if current_question_pattern is None:
                print(f"警告: row {row_no} の answer に対応する select が見つかりません")
                stats["warnings"] += 1
                continue

            old_answers = parse_answer_numbers(ws.cell(row_no, 2).value)
            if not old_answers:
                print(
                    f"警告: row {row_no} の answer は数値解答ではないため変換しません: "
                    f"{ws.cell(row_no, 2).value}"
                )
                stats["warnings"] += 1
                continue

            new_answers = convert_answers(old_answers, current_question_pattern)
            ws.cell(row_no, 7).value = format_answer_numbers(new_answers)
            stats["answer_rows"] += 1
            continue

        # ----------------------------
        # subanswer
        # ----------------------------
        if tag == "subanswer":
            if not in_subanswer_block or not subanswer_is_select:
                continue

            if current_subquest_pattern is None:
                print(f"警告: row {row_no} の subanswer に対応する subselect が見つかりません")
                stats["warnings"] += 1
                continue

            old_answers = parse_answer_numbers(ws.cell(row_no, 2).value)
            if not old_answers:
                print(
                    f"警告: row {row_no} の subanswer は数値解答ではないため変換しません: "
                    f"{ws.cell(row_no, 2).value}"
                )
                stats["warnings"] += 1
                continue

            new_answers = convert_answers(old_answers, current_subquest_pattern)
            ws.cell(row_no, 7).value = format_answer_numbers(new_answers)
            stats["subanswer_rows"] += 1
            continue

    # 最後に残っている未処理行を処理
    flush_select_rows(ws.max_row)
    flush_subselect_rows(ws.max_row)

    return stats


def resolve_excel_path(subject: str, fsyear: str, excel: str | None) -> Path:
    """
    Excelパスを決定する。
    --excel が指定されていればそれを優先。
    なければ utils.get_exam_path() で試験問題.xlsx を探す。
    """
    if excel:
        return Path(excel).expanduser().resolve()

    excel_path, _work_dir, _exam_koma_no, _sub_folder = get_exam_path(subject, fsyear)
    return Path(excel_path)


def fill_question_ids(ws, prefix: str = "Q") -> dict[str, int]:
    """
    b_question のC列に qid を毎回セットし直す。

    - 既存のC列の値は上書きする
    - Q001, Q002, Q003, ... のように連番を入れる
    - C列 = qid
    """
    stats = {
        "question_count": 0,
        "filled_qid": 0,
    }

    for row_no in range(1, ws.max_row + 1):
        tag = norm_tag(ws.cell(row_no, 1).value)

        if tag != "b_question":
            continue

        stats["question_count"] += 1

        qid = f"{prefix}{stats['question_count']:03d}"
        ws.cell(row_no, 3).value = qid  # C列

        stats["filled_qid"] += 1

    return stats

def main() -> None:
    parser = argparse.ArgumentParser(
        description="試験問題.xlsx のG列にB版用シャッフル番号と変換後正解番号をセットします。"
    )
    parser.add_argument("subject", help="科目番号。例: 1020701")
    parser.add_argument("--fsyear", default="2026", help="年度。例: 2026")
    parser.add_argument("--sheet", default=None, help="シート名。未指定なら科目番号を使います。")
    parser.add_argument("--excel", default=None, help="試験問題.xlsx のパスを直接指定する場合に使います。")
    parser.add_argument("--dry-run", action="store_true", help="保存せずに処理結果だけ確認します。")

    args = parser.parse_args()

    subject = str(args.subject)
    sheet_name = args.sheet or subject
    excel_path = resolve_excel_path(subject, args.fsyear, args.excel)

    if not excel_path.exists():
        raise FileNotFoundError(f"Excelファイルが見つかりません: {excel_path}")

    print(f"Excel: {excel_path}")
    print(f"Sheet: {sheet_name}")

    wb = load_workbook(excel_path)

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"シートが見つかりません: {sheet_name}")

    ws = wb[sheet_name]

    cleared = clear_g_column(ws)
    stats = fill_shuffle_for_sheet(ws)

    print(f"G列クリア: {cleared} セル")
    print(f"大問 select グループ: {stats['question_choice_groups']}")
    print(f"小問 subselect グループ: {stats['subquestion_choice_groups']}")
    print(f"select/subselect G列セット: {stats['choice_rows']} 行")
    print(f"answer G列セット: {stats['answer_rows']} 行")
    print(f"subanswer G列セット: {stats['subanswer_rows']} 行")
    print(f"警告: {stats['warnings']} 件")

    qid_stats = fill_question_ids(ws)

    cleared = clear_g_column(ws)
    stats = fill_shuffle_for_sheet(ws)

    print(f"b_question 数: {qid_stats['question_count']}")
    print(f"qid セット: {qid_stats['filled_qid']} 件")


    if args.dry_run:
        print("dry-run のため保存しません。")
        return

    try:
        wb.save(excel_path)
    except Exception as e:
        raise RuntimeError(
            "Excelファイルを保存できませんでした。\n"
            "Excelで開いている場合は閉じてから再実行してください。\n"
            f"対象ファイル: {excel_path}\n"
            f"元のエラー: {e}"
        )

    print("保存しました。")


if __name__ == "__main__":
    main()