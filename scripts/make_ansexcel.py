#!/usr/bin/env python3
# make_ansexcel.py — JSONデータからExcel解答用紙（学生提出用・完全指定レイアウト版）を生成
from __future__ import annotations

import json
import sys
from pathlib import Path

# Excel操作用ライブラリ
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("❌ openpyxl がインストールされていません。'pip install openpyxl' を実行してください。", file=sys.stderr)
    sys.path.append("/Users/michikazuokai/.pyenv/versions/anaconda3-2022.05/lib/python3.9/site-packages")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 既存の安全な共通処理をインポート
from exam_utils import load_exam_context, add_subject_arg

def generate_excel_sheet(json_path: Path, output_excel_path: Path, target_version: str = "A"):
    """
    指定された解答用JSONから、指定レイアウト（配点を最終行へ配置）に準拠したExcel解答用紙を生成する。
    """
    if not json_path.exists():
        raise FileNotFoundError(f"ターゲットのJSONファイルが見つかりません: {json_path}")

    with open(json_path, "r", encoding="utf-8") as f:
        outjson = json.load(f)

    # 指定されたバージョンの問題データを抽出
    version_data = None
    for v in outjson.get("versions", []):
        if v.get("version") == target_version:
            version_data = v["questions"]
            break
            
    if not version_data:
        version_data = outjson["versions"][0]["questions"]
        target_version = outjson["versions"][0]["version"]

    # 1. メタ情報（ヘッダー用）の取得
    meta_info = version_data[0]
    subject_no = meta_info.get("subject", "不明")
    title = meta_info.get("title", "試験問題")
    fsyear = meta_info.get("fsyear", "")
    kaito_message = meta_info.get("kaito_message", "")
    
    # JSONのメタ情報から年次（nenji）を取得
    nenji_raw = str(meta_info.get("nenji", "")).strip()
    nenji_text = f"{nenji_raw}年" if nenji_raw.isdigit() else nenji_raw

    # 2. ワークブックの作成と全体デザイン設定
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"解答用紙_{target_version}"
    ws.views.sheetView[0].showGridLines = True  # グリッド線を強制表示

    # スタイル定義群
    font_title = Font(name="游ゴシック", size=16, bold=True)
    font_header = Font(name="游ゴシック", size=11, bold=True)
    font_label = Font(name="Arial", size=9, bold=True, color="444444")
    font_msg = Font(name="游ゴシック", size=11, bold=True, color="333333") # 配点メッセージを少し目立たせる
    
    fill_header = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    fill_label_zone = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")

    thin_side = Side(style='thin', color='BFBFBF')
    thick_bottom = Side(style='medium', color='000000')
    
    border_cell = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    border_info = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thick_bottom)

    # 3. タイトル・受験者情報欄（ヘッダー）の組み立て
    # タイトルの末尾を (A) (B) 形式で出力
    ws["A1"] = f"{fsyear}年度  {title}  解答用紙 ({target_version})"
    ws["A1"].font = font_title
    
    # 💡 修正点: 2行目（A2）への配点メッセージ出力はここから撤廃しました（hash表示も引き続き非表示）

    # 4行目: 受験者情報欄の個別仕様化
    for col_idx in range(1, 8):
        ws.cell(row=4, column=col_idx).border = border_info

    # A列: 年次表示
    cell_a = ws.cell(row=4, column=1, value=nenji_text)
    cell_a.font = font_header
    cell_a.alignment = Alignment(horizontal="center", vertical="center")

    # B列: 「学籍」
    cell_b = ws.cell(row=4, column=2, value="学籍")
    cell_b.font = font_header
    cell_b.fill = fill_header
    cell_b.alignment = Alignment(horizontal="center", vertical="center")

    # C列: 学籍入力枠 (1セル)
    cell_c = ws.cell(row=4, column=3)
    cell_c.alignment = Alignment(horizontal="center", vertical="center")

    # D列: 「氏名」
    cell_d = ws.cell(row=4, column=4, value="氏名")
    cell_d.font = font_header
    cell_d.fill = fill_header
    cell_d.alignment = Alignment(horizontal="center", vertical="center")

    # E〜G列: 氏名入力枠として連結
    ws.merge_cells(start_row=4, start_column=5, end_row=4, end_column=7)
    ws.cell(row=4, column=5).alignment = Alignment(horizontal="left", vertical="center")

    # H列: 「点」
    cell_h = ws.cell(row=4, column=8, value="点")
    cell_h.font = font_header
    cell_h.fill = fill_header
    cell_h.alignment = Alignment(horizontal="center", vertical="center")
    cell_h.border = border_info

    # I列: 得点入力エリア
    cell_i = ws.cell(row=4, column=9)
    cell_i.alignment = Alignment(horizontal="center", vertical="center")
    cell_i.border = border_info

    ws.row_dimensions[4].height = 25

    # 4. JSONベースの解答グリッド自動生成 (6行目から開始)
    start_row = 6
    
    for block in version_data[1:]:
        if not isinstance(block, dict) or "label" not in block:
            continue

        labels = block.get("label", [])
        widths = block.get("width", [])
        heights = block.get("height", [])

        current_col = 1
        max_block_height = 1

        for i in range(len(labels)):
            label_text = str(labels[i]).strip()
            
            w_ratio = max(1, int(widths[i])) if i < len(widths) else 1
            h_ratio = max(1, int(heights[i]) // 60) if i < len(heights) else 1
            
            if h_ratio > max_block_height:
                max_block_height = h_ratio

            end_col = current_col + w_ratio - 1
            end_row = start_row + h_ratio - 1

            # セルを結合して解答マスメを作成
            ws.merge_cells(start_row=start_row, start_column=current_col, end_row=end_row, end_column=end_col)

            main_cell = ws.cell(row=start_row, column=current_col)
            main_cell.value = label_text
            main_cell.font = font_label
            main_cell.fill = fill_label_zone
            main_cell.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)

            # 結合範囲全体に外枠線を適用
            for r in range(start_row, end_row + 1):
                for c in range(current_col, end_col + 1):
                    ws.cell(row=r, column=c).border = border_cell

            current_col = end_col + 1

        # 1マスあたりの高さを設定 (38pt)
        for r in range(start_row, start_row + max_block_height):
            ws.row_dimensions[r].height = 38
            
        start_row += max_block_height + 1

    # 💡 修正点: 解答欄がすべて終わった後、1行あけて「配点：xxxx」を表示
    if kaito_message:
        # ループ終了後の start_row は、すでに前の解答ブロックから1行空いた状態を指しています
        msg_cell = ws.cell(row=start_row, column=1, value=kaito_message)
        msg_cell.font = font_msg
        msg_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[start_row].height = 20

    # 列幅の設定（A〜I列の最適化）
    ws.column_dimensions['A'].width = 8   # 年次用
    ws.column_dimensions['B'].width = 8   # 「学籍」
    ws.column_dimensions['C'].width = 14  # 学籍番号入力用
    ws.column_dimensions['D'].width = 8   # 「氏名」
    ws.column_dimensions['E'].width = 12  # 氏名枠
    ws.column_dimensions['F'].width = 12  # 氏名枠
    ws.column_dimensions['G'].width = 12  # 氏名枠
    ws.column_dimensions['H'].width = 8   # 「点」
    ws.column_dimensions['I'].width = 10  # 点数入力用

    # 解答エリア全体の列幅を一律調整
    for col_idx in range(10, max(current_col + 3, 16)):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 13

    # 5. 成果物の保存
    output_excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_excel_path)
    print(f"🎯 Excel解答用紙の生成に成功しました:\n   {output_excel_path}")


def main():
    import argparse
    parser = argparse.ArgumentParser(description="JSON から Excel 解答用紙を別個生成するスクリプト")
    add_subject_arg(parser)
    args = parser.parse_args()

    # 現行のコンテキスト特定処理
    exam_context = load_exam_context(args.subject)
    
    # JSONパスの組み立て
    json_path = exam_context.work_dir / f"{exam_context.subject}_ans.json"
    
    if not json_path.exists():
        print(f"❌ 基準データとなる JSON が見つかりません。先に make_anspdf.py を実行してください。\nパス: {json_path}", file=sys.stderr)
        sys.exit(1)

    excel_output_dir = exam_context.exam_dir / "wordexcel"
    
    with open(json_path, "r", encoding="utf-8") as f:
        outjson = json.load(f)
        
    for v_entry in outjson.get("versions", []):
        ver = v_entry.get("version", "A")
        output_file = excel_output_dir / f"{exam_context.subject}_{ver}_解答用紙.xlsx"
        
        generate_excel_sheet(json_path, output_file, target_version=ver)

if __name__ == "__main__":
    main()