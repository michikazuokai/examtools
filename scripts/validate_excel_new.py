# scripts/validate_excel.py
from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, Font

from utils import calc_excel_hash, get_exam_path

from datetime import datetime

def write_validate_log(
    work_dir: Path,
    subject: str,
    *,
    excel_path: Path,
    sheet_name: str,
    qpattern: str,
    dryrun: bool,
    excel_hash: str | None = None,
    errors: list[dict[str, Any]],
    stats: dict[str, Any] | list[str] | None = None,
    point_summary: list[str] | None = None,
) -> Path:
    """
    validate_excel.py の実行結果をログファイルに出力する。
    dry-run でも通常実行でも必ず出力する。
    """
    work_dir.mkdir(parents=True, exist_ok=True)

    log_path = work_dir / f"validate_excel_{subject}.log"

    lines: list[str] = []
    lines.append("=== validate_excel log ===")
    lines.append(f"datetime: {datetime.now().isoformat(timespec='seconds')}")
    lines.append(f"subject: {subject}")
    lines.append(f"sheet: {sheet_name}")
    lines.append(f"qpattern: {qpattern}")
    lines.append(f"excel: {excel_path}")
    lines.append(f"hash: {excel_hash or 'N/A'}")
    lines.append(f"dryrun: {dryrun}")
    lines.append("")

    lines.append("=== stats ===")
    if isinstance(stats, dict):
        for k, v in stats.items():
            lines.append(f"{k}: {v}")
    elif isinstance(stats, list):
        for line in stats:
            lines.append(str(line))
    elif stats is None:
        lines.append("No stats.")
    else:
        lines.append(str(stats))
    lines.append("")

    lines.append("=== point summary ===")
    if point_summary:
        for line in point_summary:
            lines.append(str(line))
    else:
        lines.append("No point summary.")
    lines.append("")

    lines.append("=== errors ===")
    if errors:
        for e in errors:
            row = e.get("row", "?")

            # エラー文のキー名が違っても拾えるようにする
            msg = (
                e.get("message")
                or e.get("msg")
                or e.get("error")
                or str(e)
            )

            lines.append(f"Row {row}: {msg}")
    else:
        lines.append("No errors.")
    lines.append("")

    log_path.write_text("\n".join(lines), encoding="utf-8")
    return log_path

# ============================================================
# B版選択肢シャッフル設定
# pattern[i] は「元の i+1 番の選択肢が、新しい何番へ移動するか」
# 例: [2,3,4,1] = 元1→新2, 元2→新3, 元3→新4, 元4→新1
# ============================================================
PATTERNS = [
    [2, 3, 4, 1],
    [3, 4, 1, 2],
]

TARGET_CLEAR_TAGS = {"select", "subselect", "answer", "subanswer"}

# ============================================================
# タグ構造ルール
# ============================================================
ALLOWED_CHILDREN = {
    "b_exam": {
        "examtitle", "b_examnote", "subject", "fsyear", "ansnote", "anssize",
        "b_question", "qpattern", "b_premise",
    },
    "b_examnote": {"examnote"},
    "b_premise": {"b_preline", "preimage"},
    # 旧案 pretext / 表記ゆれ preline の両方を許容
    "b_preline": {"pretext", "preline"},
    "b_question": {
        "question", "image", "sline", "b_multiline", "b_select", "b_code",
        "b_subgroup", "b_answer",
    },
    "b_subgroup": {"b_subquest"},
    "b_subquest": {
        "subquest", "subimage", "subsline", "b_submultiline", "b_subselect",
        "b_subcode", "b_subanswer",
    },
    "b_multiline": {"text"},
    "b_submultiline": {"subtext"},
    "b_select": {"select"},
    "b_subselect": {"subselect"},
    "b_code": {"code"},
    "b_subcode": {"subcode"},
    "b_answer": {"answer"},
    "b_subanswer": {"subanswer"},
}

CLOSING_TAGS = {
    "b_exam": "e_exam",
    "b_examnote": "e_examnote",
    "b_premise": "e_premise",
    "b_preline": "e_preline",
    "b_question": "e_question",
    "b_subgroup": "e_subgroup",
    "b_subquest": "e_subquest",
    "b_multiline": "e_multiline",
    "b_submultiline": "e_submultiline",
    "b_select": "e_select",
    "b_subselect": "e_subselect",
    "b_code": "e_code",
    "b_subcode": "e_subcode",
    "b_answer": "e_answer",
    "b_subanswer": "e_subanswer",
}

REQUIRED_CHILDREN = {
    "b_exam": {"examtitle", "subject", "qpattern"},
    "b_question": {"question"},
    "b_subquest": {"subquest"},
}

VALIDATION_RULES = {
    "b_select": {
        1: {
            "required": True,
            "validate": lambda val: val in {"normal", "inline"} or re.match(r"^inline\(\d+\)$", val),
            "message": "B列には normal / inline / inline(数値) のいずれかを指定してください。",
        }
    },
    "b_subselect": {
        1: {
            "required": True,
            "validate": lambda val: val in {"normal", "inline"} or re.match(r"^inline\(\d+\)$", val),
            "message": "B列には normal / inline / inline(数値) のいずれかを指定してください。",
        }
    },
    "image": {
        1: {
            "required": True,
            "validate": lambda val: re.match(r"^.+\.png(\[\d+(\.\d+)?\])?$", val),
            "message": "B列には image.png または image.png[0.8] の形式で指定してください。",
        }
    },
    "preimage": {
        1: {
            "required": True,
            "validate": lambda val: re.match(r"^.+\.png(\[\d+(\.\d+)?\])?$", val),
            "message": "B列には image.png または image.png[0.8] の形式で指定してください。",
        }
    },
}

# ============================================================
# 共通ユーティリティ
# ============================================================
def norm_tag(value: Any) -> str:
    return "" if value is None else str(value).strip()


def is_comment(tag: str) -> bool:
    if not tag:
        return False
    tag = str(tag).strip().lower()
    if tag.startswith("#") or tag.startswith("/"):
        return True
    if tag in {"コメント", "comment"}:
        return True
    return False


def is_2digits(s: Any) -> bool:
    if isinstance(s, int):
        return 1 <= s <= 99
    s = str(s).strip()
    if len(s) == 1 and s.isdigit():
        return True
    pattern = r"^[0-9]{1,2}(\s*,\s*[0-9]{1,2})*$"
    return re.match(pattern, s) is not None


def parse_answer_numbers(value: Any) -> list[int]:
    """
    answer / subanswer のB列にある元の正解番号を list[int] にする。
    2 / "2" / "1,3" / "1、3" / "1 3" に対応。
    数値以外が混じる場合は [] を返す。
    """
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []

    trans = str.maketrans("１２３４５６７８９０", "1234567890")
    text = text.translate(trans)
    text = text.replace("，", ",").replace("、", ",")
    text = re.sub(r"\s+", ",", text)

    parts = [p.strip() for p in text.split(",") if p.strip()]
    nums: list[int] = []
    for p in parts:
        if not re.fullmatch(r"\d+", p):
            return []
        n = int(p)
        if n < 1 or n > 4:
            return []
        nums.append(n)
    return nums


def convert_answers(old_answers: list[int], pattern: list[int]) -> list[int]:
    return sorted(pattern[a - 1] for a in old_answers)


def format_answer_numbers(values: list[int]) -> Any:
    values = sorted(values)
    if len(values) == 1:
        return values[0]
    return ",".join(str(v) for v in values)


def add_error(errors: list[dict[str, Any]], row: int, message: str) -> None:
    errors.append({"row": row, "message": message})

# ============================================================
# Excel補正：C列 qid / G列シャッフル
# ============================================================
def fill_question_ids(ws, prefix: str = "Q") -> dict[str, int]:
    """
    b_question のC列に qid を毎回セットし直す。
    既存値は上書きする。
    """
    stats = {"question_count": 0, "filled_qid": 0}

    for row_no in range(1, ws.max_row + 1):
        tag = norm_tag(ws.cell(row_no, 1).value)
        if tag != "b_question":
            continue
        stats["question_count"] += 1
        qid = f"{prefix}{stats['question_count']:03d}"
        ws.cell(row_no, 3).value = qid
        stats["filled_qid"] += 1

    return stats


def clear_g_column(ws) -> int:
    """
    select / subselect / answer / subanswer のG列だけをクリアする。
    b_question のG列は PB_B_after の可能性があるため消さない。
    """
    count = 0
    for row_no in range(1, ws.max_row + 1):
        tag = norm_tag(ws.cell(row_no, 1).value)
        if tag in TARGET_CLEAR_TAGS:
            if ws.cell(row_no, 7).value is not None:
                count += 1
            ws.cell(row_no, 7).value = None
    return count


def apply_pattern_to_choice_rows(ws, rows: list[int], pattern: list[int]) -> None:
    if len(rows) != 4:
        raise ValueError(f"選択肢が4行ではありません: rows={rows}")
    for row_no, value in zip(rows, pattern):
        ws.cell(row_no, 7).value = value


def fill_shuffle_for_sheet(ws) -> dict[str, int]:
    """
    b_question〜e_question、b_subquest〜e_subquest の範囲を意識して、
    G列にB版用シャッフル番号と変換後正解番号をセットする。
    """
    stats = {
        "question_choice_groups": 0,
        "subquestion_choice_groups": 0,
        "choice_rows": 0,
        "answer_rows": 0,
        "subanswer_rows": 0,
        "warnings": 0,
    }

    in_question = False
    in_subquest = False
    in_answer_block = False
    in_subanswer_block = False
    answer_is_select = False
    subanswer_is_select = False

    current_question_pattern: list[int] | None = None
    current_subquest_pattern: list[int] | None = None
    select_rows: list[int] = []
    subselect_rows: list[int] = []
    question_select_group_index = 0
    subquestion_select_group_index = 0

    def warn(msg: str) -> None:
        print(f"警告: {msg}")
        stats["warnings"] += 1

    def flush_select_rows(current_row: int) -> None:
        nonlocal select_rows, current_question_pattern, question_select_group_index
        if not select_rows:
            return
        if len(select_rows) != 4:
            warn(f"select が4行ではありません: rows={select_rows}, near row={current_row}")
            select_rows = []
            current_question_pattern = None
            return
        pattern = PATTERNS[question_select_group_index % len(PATTERNS)]
        apply_pattern_to_choice_rows(ws, select_rows, pattern)
        current_question_pattern = pattern
        question_select_group_index += 1
        stats["question_choice_groups"] += 1
        stats["choice_rows"] += 4
        select_rows = []

    def flush_subselect_rows(current_row: int) -> None:
        nonlocal subselect_rows, current_subquest_pattern, subquestion_select_group_index
        if not subselect_rows:
            return
        if len(subselect_rows) != 4:
            warn(f"subselect が4行ではありません: rows={subselect_rows}, near row={current_row}")
            subselect_rows = []
            current_subquest_pattern = None
            return
        pattern = PATTERNS[subquestion_select_group_index % len(PATTERNS)]
        apply_pattern_to_choice_rows(ws, subselect_rows, pattern)
        current_subquest_pattern = pattern
        subquestion_select_group_index += 1
        stats["subquestion_choice_groups"] += 1
        stats["choice_rows"] += 4
        subselect_rows = []

    for row_no in range(1, ws.max_row + 1):
        tag = norm_tag(ws.cell(row_no, 1).value)

        if tag == "b_question":
            flush_select_rows(row_no)
            flush_subselect_rows(row_no)
            in_question = True
            current_question_pattern = None
            select_rows = []
            continue

        if tag == "e_question":
            flush_select_rows(row_no)
            flush_subselect_rows(row_no)
            in_question = False
            in_answer_block = False
            answer_is_select = False
            current_question_pattern = None
            select_rows = []
            continue

        if tag == "b_subquest":
            flush_select_rows(row_no)
            flush_subselect_rows(row_no)
            in_subquest = True
            current_subquest_pattern = None
            subselect_rows = []
            continue

        if tag == "e_subquest":
            flush_subselect_rows(row_no)
            in_subquest = False
            in_subanswer_block = False
            subanswer_is_select = False
            current_subquest_pattern = None
            subselect_rows = []
            continue

        if tag == "select":
            if not in_question:
                warn(f"row {row_no} の select が b_question〜e_question の外にあります")
                continue
            select_rows.append(row_no)
            if len(select_rows) == 4:
                flush_select_rows(row_no)
            continue

        if select_rows and tag != "select":
            flush_select_rows(row_no)

        if tag == "subselect":
            if not in_subquest:
                warn(f"row {row_no} の subselect が b_subquest〜e_subquest の外にあります")
                continue
            subselect_rows.append(row_no)
            if len(subselect_rows) == 4:
                flush_subselect_rows(row_no)
            continue

        if subselect_rows and tag != "subselect":
            flush_subselect_rows(row_no)

        if tag == "b_answer":
            in_answer_block = True
            marker = ws.cell(row_no, 2).value
            answer_is_select = str(marker).strip() == "#select" if marker is not None else False
            continue

        if tag == "e_answer":
            in_answer_block = False
            answer_is_select = False
            continue

        if tag == "b_subanswer":
            in_subanswer_block = True
            marker = ws.cell(row_no, 2).value
            subanswer_is_select = str(marker).strip() == "#select" if marker is not None else False
            continue

        if tag == "e_subanswer":
            in_subanswer_block = False
            subanswer_is_select = False
            continue

        if tag == "answer":
            if not in_answer_block or not answer_is_select:
                continue
            if current_question_pattern is None:
                warn(f"row {row_no} の answer に対応する select が見つかりません")
                continue
            old_answers = parse_answer_numbers(ws.cell(row_no, 2).value)
            if not old_answers:
                warn(f"row {row_no} の answer は数値解答ではないため変換しません: {ws.cell(row_no, 2).value}")
                continue
            new_answers = convert_answers(old_answers, current_question_pattern)
            ws.cell(row_no, 7).value = format_answer_numbers(new_answers)
            stats["answer_rows"] += 1
            continue

        if tag == "subanswer":
            if not in_subanswer_block or not subanswer_is_select:
                continue
            if current_subquest_pattern is None:
                warn(f"row {row_no} の subanswer に対応する subselect が見つかりません")
                continue
            old_answers = parse_answer_numbers(ws.cell(row_no, 2).value)
            if not old_answers:
                warn(f"row {row_no} の subanswer は数値解答ではないため変換しません: {ws.cell(row_no, 2).value}")
                continue
            new_answers = convert_answers(old_answers, current_subquest_pattern)
            ws.cell(row_no, 7).value = format_answer_numbers(new_answers)
            stats["subanswer_rows"] += 1
            continue

    flush_select_rows(ws.max_row)
    flush_subselect_rows(ws.max_row)
    return stats

# ============================================================
# 構造・値チェック
# ============================================================
def check_structure(tag: str, stack: list[str], rownum: int, errors: list[dict[str, Any]], block_children: dict[str, list[set[str]]]) -> None:
    if tag.startswith("b_"):
        stack.append(tag)
        block_children.setdefault(tag, []).append(set())
        return

    if tag.startswith("e_"):
        if not stack:
            add_error(errors, rownum, f"'{tag}' に対応する開始タグがありません。")
            return
        expected = "b_" + tag[2:]
        if stack[-1] != expected:
            add_error(errors, rownum, f"'{tag}' が現在開いている '{stack[-1]}' と対応していません。")
            return

        block_tag = stack[-1]
        seen = block_children.get(block_tag, [set()])[-1]

        # 通常の必須子タグチェック
        required = REQUIRED_CHILDREN.get(block_tag, set())
        for miss in sorted(required - seen):
            add_error(errors, rownum, f"'{block_tag}' の中に必須タグ '{miss}' がありません。")

        # b_preline は pretext / preline のどちらかがあればOK
        if block_tag == "b_preline" and not ({"pretext", "preline"} & seen):
            add_error(errors, rownum, "'b_preline' の中に pretext または preline がありません。")

        stack.pop()
        if block_tag in block_children and block_children[block_tag]:
            block_children[block_tag].pop()
        return

    if not stack:
        add_error(errors, rownum, f"'{tag}' がどのブロックの中にもありません。")
        return

    parent = stack[-1]
    allowed = ALLOWED_CHILDREN.get(parent, set())
    if tag not in allowed:
        add_error(errors, rownum, f"'{tag}' は '{parent}' の中では使用できません。")
        return

    if parent not in block_children or not block_children[parent]:
        block_children.setdefault(parent, []).append(set())
    block_children[parent][-1].add(tag)


def check_values(tag: str, row: tuple[Any, ...], rownum: int, errors: list[dict[str, Any]], state: dict[str, Any]) -> None:
    if tag in {"b_answer", "b_subanswer"}:
        state["check_answer"] = False
        try:
            state["check_answer"] = str(row[1]).strip().lower() == "#select"
        except Exception:
            state["check_answer"] = False
        return

    if tag in {"e_answer", "e_subanswer"}:
        state["check_answer"] = False
        return

    if tag in {"answer", "subanswer"}:
        if state.get("check_answer"):
            if len(row) < 2 or not is_2digits(row[1]):
                add_error(errors, rownum, f"'{tag}' のB列は 1〜2桁の数値、またはカンマ区切りの数値にしてください。")
        return

    rule = VALIDATION_RULES.get(tag)
    if not rule:
        return

    for col_index, rule_item in rule.items():
        required = rule_item.get("required", False)
        validator = rule_item.get("validate")
        message = rule_item.get("message", f"列{col_index + 1}の値が不正です。")

        if len(row) <= col_index or row[col_index] is None:
            if required:
                add_error(errors, rownum, f"'{tag}' の{col_index + 1}列目が未入力です。{message}")
            continue

        cell_value = str(row[col_index]).strip()
        if not cell_value:
            if required:
                add_error(errors, rownum, f"'{tag}' の{col_index + 1}列目が空です。{message}")
            continue

        if validator and not validator(cell_value):
            add_error(errors, rownum, f"'{tag}' の{col_index + 1}列目の値 '{cell_value}' が不正です。{message}")

def validate_question_columns(ws, errors: list[dict[str, Any]], make_b: bool = False) -> None:
    """
    b_question のC列 qid / D列 orderB をチェックする。

    qid:
        A版/B版に関係なく必要なので常にチェックする。

    orderB:
        B版を作る場合だけ必要なので、make_b=True のときだけチェックする。
    """
    qids: list[str] = []
    orders: list[int] = []

    for row_no in range(1, ws.max_row + 1):
        tag = norm_tag(ws.cell(row_no, 1).value)
        if tag != "b_question":
            continue

        # C列 qid は常にチェック
        qid = ws.cell(row_no, 3).value
        if qid is None or str(qid).strip() == "":
            add_error(errors, row_no, "b_question のC列(qid)が未入力です。validate_excel.pyで自動セットできます。")
        else:
            qids.append(str(qid).strip())

        # D列 orderB はB版を作る場合だけチェック
        if make_b:
            orderB = ws.cell(row_no, 4).value
            if orderB is None or str(orderB).strip() == "":
                add_error(errors, row_no, "b_question のD列(orderB)が未入力です。B版を作る場合は並び順を入力してください。")
            else:
                try:
                    orders.append(int(str(orderB).strip()))
                except Exception:
                    add_error(errors, row_no, f"b_question のD列(orderB)は整数にしてください: {orderB}")

    # qid の重複は常にチェック
    dup_qids = sorted({x for x in qids if qids.count(x) > 1})
    for qid in dup_qids:
        add_error(errors, 1, f"qid が重複しています: {qid}")

    # orderB の重複はB版を作る場合だけチェック
    if make_b:
        dup_orders = sorted({x for x in orders if orders.count(x) > 1})
        for ob in dup_orders:
            add_error(errors, 1, f"orderB が重複しています: {ob}")


def __validate_question_columns(ws, errors: list[dict[str, Any]]) -> None:
    """b_question のC列 qid / D列 orderB をチェックする。"""
    qids: list[str] = []
    orders: list[int] = []

    for row_no in range(1, ws.max_row + 1):
        tag = norm_tag(ws.cell(row_no, 1).value)
        if tag != "b_question":
            continue

        qid = ws.cell(row_no, 3).value
        if qid is None or str(qid).strip() == "":
            add_error(errors, row_no, "b_question のC列(qid)が未入力です。--fix で自動セットできます。")
        else:
            qids.append(str(qid).strip())

        orderB = ws.cell(row_no, 4).value
        if orderB is None or str(orderB).strip() == "":
            add_error(errors, row_no, "b_question のD列(orderB)が未入力です。B版を作る場合は並び順を入力してください。")
        else:
            try:
                orders.append(int(str(orderB).strip()))
            except Exception:
                add_error(errors, row_no, f"b_question のD列(orderB)は整数にしてください: {orderB}")

    dup_qids = sorted({x for x in qids if qids.count(x) > 1})
    for qid in dup_qids:
        add_error(errors, 1, f"qid が重複しています: {qid}")

    dup_orders = sorted({x for x in orders if orders.count(x) > 1})
    for ob in dup_orders:
        add_error(errors, 1, f"orderB が重複しています: {ob}")


def validate_special_rows(ws, errors: list[dict[str, Any]]) -> None:
    """LINESPACE / PAGEBREAK など、構造スタックだけでは見にくいルール。"""
    in_subgroup = False
    multiline_count = 0
    submultiline_count = 0
    in_multiline = False
    in_submultiline = False

    for row_no in range(1, ws.max_row + 1):
        tag = norm_tag(ws.cell(row_no, 1).value)
        if not tag or is_comment(tag) or "#" in tag:
            continue

        if tag == "b_subgroup":
            in_subgroup = True
        elif tag == "e_subgroup":
            in_subgroup = False

        if tag == "b_multiline":
            in_multiline = True
            multiline_count = 0
        elif tag == "text" and in_multiline:
            multiline_count += 1
        elif tag == "e_multiline":
            if multiline_count == 0:
                add_error(errors, row_no, "b_multiline〜e_multiline の中に text がありません。")
            in_multiline = False

        if tag == "b_submultiline":
            in_submultiline = True
            submultiline_count = 0
        elif tag == "subtext" and in_submultiline:
            submultiline_count += 1
        elif tag == "e_submultiline":
            if submultiline_count == 0:
                add_error(errors, row_no, "b_submultiline〜e_submultiline の中に subtext がありません。")
            in_submultiline = False

        if tag == "PAGEBREAK":
            add_error(errors, row_no, "PAGEBREAK 行タグは使用しません。b_question の E/G列(PB_*_after)を使ってください。")

        if tag == "LINESPACE":
            if not in_subgroup:
                add_error(errors, row_no, "LINESPACE は b_subgroup〜e_subgroup の中で使用してください。")
            v = ws.cell(row_no, 2).value
            try:
                float(v)
            except Exception:
                add_error(errors, row_no, "LINESPACE のB列には数値を入力してください。")

def get_qpattern(ws) -> str:
    """
    qpattern タグのB列から A / A,B / B を取得する。
    見つからなければ A とみなす。
    """
    for row_no in range(1, ws.max_row + 1):
        tag = norm_tag(ws.cell(row_no, 1).value)
        if tag == "qpattern":
            value = ws.cell(row_no, 2).value
            return str(value).strip().upper() if value else "A"

    return "A"

def validate_sheet(ws) -> list[dict[str, Any]]:
    errors: list[dict[str, Any]] = []
    stack: list[str] = []
    block_children: dict[str, list[set[str]]] = {}
    value_state = {"check_answer": False}

    for rownum, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if not row or not row[0]:
            continue
        tag = norm_tag(row[0])
        if is_comment(tag) or "#" in tag:
            continue
        if tag in {"PAGEBREAK", "LINESPACE"}:
            continue
        check_structure(tag, stack, rownum, errors, block_children)
        check_values(tag, row, rownum, errors, value_state)

    if stack:
        add_error(errors, ws.max_row, f"閉じられていないブロックがあります: {stack}")

    qpattern = get_qpattern(ws)
    make_b = "B" in qpattern
    validate_question_columns(ws, errors, make_b=make_b)

    validate_special_rows(ws, errors)
    return errors

# ============================================================
# 点数集計
# ============================================================
def tokutenlst(ws) -> list[str]:
    i = 1
    n = ws.max_row
    selflg = False
    totten = 0
    outlst: list[str] = []
    svcode = ""

    while i <= n:
        code = ws.cell(row=i, column=1).value
        ten = ws.cell(row=i, column=3).value

        if code is None:
            i += 1
            continue

        if str(code).strip().startswith("# 【前提条件】"):
            i += 1
            continue

        if "【" in str(code) or "問題" in str(code) or "問" in str(code):
            text = str(code)
            m = re.search(r"(?:問|問題)\s*(\d+)(?:-(\d+))?", text)
            if m:
                big = m.group(1)
                sub = m.group(2)
                svcode = f"問{big}-{sub}" if sub else f"問{big}"
            else:
                svcode = str(code)

        if code == "b_question":
            totqten = 0
            selflg = True

        if selflg and code == "e_question":
            # 小問がない大問用。小問がある場合は大問自体の点が0なら出ても害は小さい。
            if totqten:
                outlst.append(f"{svcode}    {totqten}")
                totten += totqten

        if selflg and code == "answer":
            try:
                totqten += int(ten or 0)
            except Exception:
                pass

        if code == "b_subquest":
            totsubten = 0
            selflg = True

        if selflg and code == "subanswer":
            try:
                totsubten += int(ten or 0)
            except Exception:
                pass

        if selflg and code == "e_subquest":
            outlst.append(f"{svcode}  {totsubten}")
            totten += totsubten
            selflg = False

        i += 1

    outlst.append(f"合計  {totten}")
    return outlst

# ============================================================
# コメント・stamp
# ============================================================
def clear_validation_comments(ws) -> int:
    """A列の validate_excel コメントだけを削除する。"""
    count = 0
    for row_no in range(1, ws.max_row + 1):
        cell = ws.cell(row_no, 1)
        if cell.comment and "validate_excel" in str(cell.comment.text):
            cell.comment = None
            count += 1
    return count


def apply_validation_comments(ws, errors: list[dict[str, Any]]) -> int:
    grouped: dict[int, list[str]] = {}
    for e in errors:
        grouped.setdefault(int(e["row"]), []).append(str(e["message"]))

    for row_no, messages in grouped.items():
        text = "validate_excel\n" + "\n".join(f"- {m}" for m in messages)
        ws.cell(row_no, 1).comment = Comment(text, "validate_excel")

    return len(grouped)


def write_validation_stamp(excel_path: Path, work_dir: Path, subject: str, sheetname: str, ws, qpattern) -> Path:
    work_dir.mkdir(parents=True, exist_ok=True)
    stamp = {
        "subject": str(subject),
        "sheetname": str(sheetname),
        "qpattern": qpattern,
        "inputpath": str(excel_path),
        "hash": calc_excel_hash(ws),
        "status": "ok",
        "validated_by": "validate_excel.py",
        "validated_at": datetime.now().isoformat(timespec="seconds"),
    }
    stamp_path = work_dir / f"validation_stamp_{subject}.json"
    stamp_path.write_text(json.dumps(stamp, ensure_ascii=False, indent=2), encoding="utf-8")
    return stamp_path

# ============================================================
# 実行処理
# ============================================================
def resolve_excel_path(subject: str, fsyear: str, excel: str | None) -> tuple[Path, Path, str, Path]:
    if excel:
        excel_path = Path(excel).expanduser().resolve()
        return excel_path, excel_path.parent / "work", "", excel_path.parent
    return get_exam_path(subject, fsyear)

def apply_answer_styles(ws) -> dict[str, int]:
    """
    解答タグ行の見た目を整える。

    対象タグ:
      b_answer, answer, e_answer,
      b_subanswer, subanswer, e_subanswer

    処理:
      - 対象行のA列〜G列の背景色を FCE4D6 にする
      - b_answer / b_subanswer 行は、#select の有無に関係なく、
        A列〜G列の文字を
          游ゴシック / 12pt / 太字 / 0070C0
        にする
    """
    from openpyxl.styles import PatternFill, Font

    target_tags = {
        "b_answer",
        "answer",
        "e_answer",
        "b_subanswer",
        "subanswer",
        "e_subanswer",
    }

    header_tags = {"b_answer", "b_subanswer"}

    fill = PatternFill(
        fill_type="solid",
        fgColor="FFFCE4D6",
    )

    header_font = Font(
        name="游ゴシック",
        size=12,
        bold=True,
        color="FF0070C0",
    )

    stats = {
        "styled_answer_rows": 0,
        "styled_answer_header_rows": 0,
    }

    for row_no in range(1, ws.max_row + 1):
        tag = norm_tag(ws.cell(row_no, 1).value)

        if tag not in target_tags:
            continue

        # 対象タグ行は、A列〜G列に背景色を設定
        for col_no in range(1, 8):
            ws.cell(row_no, col_no).fill = fill

        stats["styled_answer_rows"] += 1

        # b_answer / b_subanswer 行は、#select の有無に関係なく文字を強調
        if tag in header_tags:
            for col_no in range(1, 8):
                ws.cell(row_no, col_no).font = header_font

            stats["styled_answer_header_rows"] += 1

    return stats

def run_validate(
    excel_path: Path,
    sheetname: str,
    *,
    save: bool,
) -> tuple[list[dict[str, Any]], list[str], dict[str, Any]]:
    """
    統合版の実行本体。

    通常実行では、次を必ず行う。
      - C列 qid を毎回セット
      - G列 shuffle / B版正解を毎回セット
      - 既存の validate_excel コメントを削除
      - エラーコメントをA列にセット
      - Excelを保存

    dryrun の場合は save=False として呼び出す。
    """
    wb = openpyxl.load_workbook(excel_path)
    if sheetname not in wb.sheetnames:
        raise ValueError(f"シートが見つかりません: {sheetname}")
    ws = wb[sheetname]

    stats: dict[str, Any] = {}

    # 毎回、前処理として値を作り直す
    stats["qid"] = fill_question_ids(ws)
    stats["g_cleared"] = clear_g_column(ws)
    stats["shuffle"] = fill_shuffle_for_sheet(ws)

    stats["answer_styles"] = apply_answer_styles(ws)

    # コメントは毎回作り直す
    stats["cleared_comments"] = clear_validation_comments(ws)

    errors = validate_sheet(ws)
    score_list = tokutenlst(ws)

    stats["error_comments"] = apply_validation_comments(ws, errors)

    if save:
        try:
            wb.save(excel_path)
        except Exception as e:
            raise RuntimeError(
                "Excelファイルを保存できませんでした。\n"
                "Excelで開いている場合は閉じてから再実行してください。\n"
                f"対象ファイル: {excel_path}\n"
                f"元のエラー: {e}"
            )
    excel_hash = calc_excel_hash(ws)
    qpattern = get_qpattern(ws)
    return errors, score_list, stats, excel_hash, qpattern


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "試験問題.xlsx をチェックし、C列qid・G列B版シャッフル・コメントを更新します。"
            "通常実行ではExcelを保存し、Validation OK時にstampを作成します。"
        )
    )
    parser.add_argument("subject", help="科目番号。例: 1020701")
    parser.add_argument("fsyear_pos", nargs="?", default=None, help="年度。例: 2026（任意）")
    parser.add_argument("sheet_pos", nargs="?", default=None, help="シート名（任意。未指定なら科目番号）")
    parser.add_argument("--fsyear", default=None, help="年度。例: 2026")
    parser.add_argument("--sheet", default=None, help="シート名。未指定なら科目番号")
    parser.add_argument("--excel", default=None, help="試験問題.xlsx のパスを直接指定")
    parser.add_argument("--dryrun", action="store_true", help="保存せずに処理結果だけ確認します。stampも作成しません。")

    args = parser.parse_args()

    subject = str(args.subject)
    fsyear = args.fsyear or args.fsyear_pos or "2026"
    sheetname = args.sheet or args.sheet_pos or subject

    excel_path, work_dir, exam_koma_no, sub_folder = resolve_excel_path(subject, fsyear, args.excel)

    print(f"科目番号: {subject}")
    print(f"年度: {fsyear}")
    print(f"シート名: {sheetname}")
    if exam_koma_no:
        print(f"試験コマ番号: {exam_koma_no}")
    print(f"入力Excel: {excel_path}")

    if not excel_path.exists():
        raise FileNotFoundError(f"Excelファイルが見つかりません: {excel_path}")

    should_save = not args.dryrun

    errors, score_list, stats, excel_hash, qpattern = run_validate(
        excel_path,
        sheetname,
        save=should_save,
    )

    qid_stats = stats.get("qid", {})
    print(f"b_question数: {qid_stats.get('question_count', 0)}")
    print(f"qidセット: {qid_stats.get('filled_qid', 0)} 件")

    print(f"G列クリア: {stats.get('g_cleared', 0)} セル")
    sh = stats.get("shuffle", {})
    print(f"大問 select グループ: {sh.get('question_choice_groups', 0)}")
    print(f"小問 subselect グループ: {sh.get('subquestion_choice_groups', 0)}")
    print(f"select/subselect G列セット: {sh.get('choice_rows', 0)} 行")
    print(f"answer G列セット: {sh.get('answer_rows', 0)} 行")
    print(f"subanswer G列セット: {sh.get('subanswer_rows', 0)} 行")
    print(f"shuffle警告: {sh.get('warnings', 0)} 件")

    print(f"既存コメント削除: {stats.get('cleared_comments', 0)} 件")
    print(f"エラーコメントセット: {stats.get('error_comments', 0)} 件")

    print("点数状況:")
    for v in score_list:
        print(v)

    # dryrun でも通常実行でも、常にログを出力する
    log_path = write_validate_log(
        Path(work_dir),
        subject,
        excel_path=Path(excel_path),
        sheet_name=sheetname,
        qpattern=qpattern,
        dryrun=args.dryrun,
        excel_hash=excel_hash,
        errors=errors,
        stats=stats,
        point_summary=score_list,
    )
    print(f"ログ出力: {log_path}")

    if errors:
        print("Validation errors:")
        for e in errors:
            print(f" - Row {e['row']}: {e['message']}")
        if should_save:
            print("エラーがあります。Excelは保存しましたが、validation stamp は作成しません。")
            print("Excelを開いてコメントを確認し、修正後に再実行してください。")
        else:
            print("dryrun のため保存していません。")
        sys.exit(1)

    print("Validation OK!")

    if should_save:
        print("Excelを保存しました。")
        # 保存後の内容でstampを作る
        wb2 = openpyxl.load_workbook(excel_path, data_only=False)
        ws2 = wb2[sheetname]
        stamp_path = write_validation_stamp(excel_path, Path(work_dir), subject, sheetname, ws2, qpattern)
        print(f"validation stamp: {stamp_path}")
    else:
        print("dryrun のため保存していません。validation stamp も作成していません。")


if __name__ == "__main__":
    main()
