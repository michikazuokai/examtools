import argparse
import openpyxl
import re

# ----------------------------
# 子タグルール定義
# ----------------------------
ALLOWED_CHILDREN = {
    "b_exam": {"examtitle", "b_examnote", "subject", "fsyear", "ansnote", "anssize",
               "b_question", "qpattern"},
    "b_examnote": {"examnote"},
    "b_question": {"question", "image", "sline", "b_multiline", 
                   "b_select", "b_code", "b_subgroup",
                   "b_answer"   # ★ 追加
                   },
    "b_subgroup": {"b_subquest"},
    "b_subquest": {"subquest", "subimage", "subsline", "b_submultiline", 
                   "b_subselect", "b_subcode",
                   "b_subanswer",   # ★ 追加
                   },
    "b_multiline": {"text"},
    "b_submultiline": {"subtext"},
    "b_select": {"select"},
    "b_subselect": {"subselect"},
    "b_code": {"code"},
    "b_subcode": {"subcode"},
    "b_answer": {"answer"},   # ★ b_answer の中身は answer のみ
    "b_subanswer": {"subanswer"},   
}

# ----------------------------
# 対応する閉じタグ
# ----------------------------
CLOSING_TAGS = {
    "b_exam": "e_exam",
    "b_examnote": "e_examnote",   # ← これを追加
    "b_question": "e_question",
    "b_subgroup": "e_subgroup",
    "b_subquest": "e_subquest",
    "b_multiline": "e_multiline",
    "b_submultiline": "e_submultiline",
    "b_select": "e_select",
    "b_subselect": "e_subselect",
    "b_code": "e_code",
    "b_subcode": "e_subcode",
    "b_answer": "e_answer",
    "b_subanswer": "e_subanswer",
}

# ----------------------------
# 必須の子タグ定義（ブロック単位）
# ----------------------------
REQUIRED_CHILDREN = {
    "b_exam": {"examtitle", "subject", "qpattern"},
    "b_question": {"question", },
    "b_subquest": {"subquest"},
}

def is_comment(tag: str) -> bool:
    """コメント行判定"""
    if not tag:
        return False
    tag = str(tag).strip().lower()
    if tag.startswith("#") or tag.startswith("/"):
        return True
    if tag in {"コメント", "comment"}:
        return True
    return False

# 整合チェックルール定義（列番号は0始まり）
def is_integer(v):
    try:
        int(v)
        return True
    except ValueError:
        return False
    
VALIDATION_RULES = {
    "b_select": {
        1: {
            "required": True,
            "validate": lambda val: val in {"normal", "inline"} or re.match(r"^inline\(\d+\)$", val)
        }
    },
    "b_subselect": {
        1: {
            "required": True,
            "validate": lambda val: val in {"normal", "inline"} or re.match(r"^inline\(\d+\)$", val)
        }
    },
     "image": {
         1: {
             "required": True,
             "validate": lambda val: re.match(r'^.+\.png(\[\d+(\.\d+)?\])?$', val)
         }
     },
     "answer": {
         1: {"required": True},
         2: {"required": True,"validate": is_integer}
    },
     "subanswer": {
         1: {"required": True},
         2: {"required": True,"validate": is_integer}
    }
}

def check_structure(tag, stack, rownum, errors):
    # ✅ ブロック内で出現した子タグを記録する辞書（関数属性として初期化）
    if not hasattr(check_structure, "block_children"):
        check_structure.block_children = {}

    # 中身タグ処理
    if tag.startswith("b_"):
        stack.append(tag)
        # ✅ 初出のときだけセットする
        if tag not in check_structure.block_children:
            check_structure.block_children[tag] = set()
    elif tag.startswith("e_"):
        if not stack:
            errors.append(f"Row {rownum}: '{tag}' without matching opening tag")
        else:
            expected = "b_" + tag[2:]
            if stack[-1] != expected:
                errors.append(f"Row {rownum}: '{tag}' does not match open tag '{stack[-1]}'")
            else:
                # ✅ 必須タグチェック（ブロック終了時）
                block_tag = stack[-1]
                required = REQUIRED_CHILDREN.get(block_tag, set())
                seen = check_structure.block_children.get(block_tag, set())
                missing = required - seen
                for miss in missing:
                    errors.append(f"Row {rownum}: '{miss}' is required inside '{block_tag}' but missing.")
                stack.pop()
                # ✅ 使用済みのブロックの子要素記録を削除
                if block_tag in check_structure.block_children:
                    del check_structure.block_children[block_tag]
    else:
        # 通常タグ（中身タグ）
        if stack:
            parent = stack[-1]
            allowed = ALLOWED_CHILDREN.get(parent, set())
            if tag not in allowed:
                errors.append(f"Row {rownum}: '{tag}' not allowed inside '{parent}'")
            else:
                # ✅ 親ブロックの記録に追加（積み重ねる）
                if parent not in check_structure.block_children:
                    check_structure.block_children[parent] = set()
                check_structure.block_children[parent].add(tag)
        else:
            errors.append(f"Row {rownum}: '{tag}' outside of any block")

def is_2digits(s):
    # 数字（int型）の場合：1-99の範囲ならOK
    if isinstance(s, int):
        return 1 <= s <= 99
    
    # 文字列の場合：既存のチェック
    s = str(s).strip()
    if len(s) == 1 and s.isdigit():
        return True
    
    pattern = r'^[0-9]{1,2}(\s*,\s*[0-9]{1,2})*$'
    return re.match(pattern, s) is not None


def check_values(tag, row, rownum, errors):
    """タグごとの値チェック。select/subselect に紐づく answer/subanswer も検証."""

    # -------------------------------------------------------
    # ① 関数属性でフラグを初期化（最初の呼び出し時）
    # -------------------------------------------------------
    if not hasattr(check_values, "check_answer"):
        check_values.check_answer = False  # answer/subanswer チェック用

    # -------------------------------------------------------
    # ② b_answer / b_subanswer が "select" を指定しているか判定
    # -------------------------------------------------------
    if tag in {"b_answer", "b_subanswer"}:
        # B列が select なら true
        try:
            check_values.check_answer = (str(row[1]).strip().lower() == "#select")
        except Exception:
            check_values.check_answer = False
        return  # 追加チェックは不要

    # -------------------------------------------------------
    # ③ answer/subanswer 本体のチェック
    # -------------------------------------------------------
    if tag in {"answer", "subanswer"}:
        if check_values.check_answer:
            # B列の形式チェック
            if len(row) < 2 or not is_2digits(row[1]):
                errors.append(
                    f"Row {rownum}: Invalid answer value '{row[1]}' for tag '{tag}' "
                    f"(must be 1–2 digits or comma-separated 1–2 digits)"
                )
        return  # 通常の VALIDATION_RULES は適用しない

    # -------------------------------------------------------
    # ③ 通常の VALIDATION_RULES によるチェック
    # -------------------------------------------------------
    rule = VALIDATION_RULES.get(tag)
    if rule:
        for col_index, rule_item in rule.items():
            required = rule_item.get("required", False)
            validator = rule_item.get("validate")

            # 列数不足 or None
            if len(row) <= col_index or row[col_index] is None:
                if required:
                    errors.append(f"Row {rownum}: Missing required column {col_index+1} for tag '{tag}'")
                continue

            cell_value = str(row[col_index]).strip()
            if not cell_value:
                if required:
                    errors.append(f"Row {rownum}: Empty value in required column {col_index+1} for tag '{tag}'")
                continue

            if validator and not validator(cell_value):
                errors.append(
                    f"Row {rownum}: Invalid value '{cell_value}' in column {col_index+1} for tag '{tag}'"
                )

def tokutenlst(ws):
    i = 1                      # openpyxl は 1 始まり
    n = ws.max_row
    selflg = False
    totten=0
    outlst=[]
    while i <= n:
        code = ws.cell(row=i, column=1).value   # 1列目 (コード)
        ten = ws.cell(row=i, column=3).value   # 得点

        if code is None:
            i += 1
            continue

        # 問題コード行の保存
        if "【" in str(code) or "問題" in str(code) or "問" in str(code):
            text = str(code)

            # 問12 または 問題12 または 問12-3 または 問題12-3 を全て処理
            m = re.search(r"(?:問|問題)\s*(\d+)(?:-(\d+))?", text)

            if m:
                big = m.group(1)      # 問題番号
                sub = m.group(2)      # 小問番号（無い場合は None）
                if sub:
                    svcode = f"問{big}-{sub}"
                else:
                    svcode = f"問{big}"
            else:
                print("番号を検出できません:", repr(text))
                svcode = str(code)

        # --- 大問 select ---
        if code == "b_question":
            totqten=0
            selflg = True

        if selflg and code == "e_question":
            outlst.append(f'{svcode}    {totqten}')
            totten+=totqten

        if selflg and code == "answer":
            totqten+=ten

        # --- 小問 subselect ---
        if code == "b_subquest":
            totsubten=0
            selflg = True

        if selflg and code == "subanswer":
            totsubten+=ten

        if selflg and code == "e_subquest":
            outlst.append(f'{svcode}  {totsubten}')
            totten+=totsubten
            selflg = False

        i += 1

    outlst.append(f"合計  {totten}")
    return outlst

def validate_excel(path: str, shname="Sheet1") -> list[str]:
    """Excelファイルを読み込み、構造チェックを行う"""
    wb = openpyxl.load_workbook(path)
    ws = wb[shname]

    errors = []
    stack = []  # 開いているブロックのスタック
    toklst = []  # ← ここで初期化しておく

    for rownum, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if not row or not row[0]:
            continue
        tag = str(row[0]).strip()

        # コメント行はスキップ
        if is_comment(tag) or "#" in tag:
            continue

        # ----------------------------
        # 中身タグ処理
        # ----------------------------
        if tag == "PAGEBREAK"or tag == "LINESPACE":
            continue  # どの親でもOK

        check_structure(tag, stack, rownum, errors)
        check_values(tag, row, rownum, errors)

    # ----------------------------
    # 最後に stack が空でなければエラー
    # ----------------------------
    if stack:
        errors.append(f"Unclosed blocks at end: {stack}")
    if not errors:
        toklst=tokutenlst(ws)
    return errors,toklst

from pathlib import Path
import sys
def main():
    # 現在の場所
    curdir = Path(__file__).parent.parent

    # 試験用のエクセルファイル デフォルトは "試験問題.xlsx"
    examdata = "試験問題.xlsx" if len(sys.argv) < 3 else sys.argv[2]

    excel_path = curdir / 'input' / examdata
    sheetname = sys.argv[1]  #シート名：科目番号

    errs,toklst = validate_excel(excel_path, sheetname)
    if errs:
        print("Validation errors:")
        for e in errs:
            print(" -", e)
            sys.exit(1) 
    else:
        print("Validation OK!")
        for v in toklst:
            print(v)
        sys.exit(0) 


if __name__ == "__main__":
    main()