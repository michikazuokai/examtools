# anstest_answer.py
# Excel(試験問題2.xlsx 等) → 解答用JSON（列単位）
from copy import deepcopy
import json
from openpyxl import load_workbook
from ansmake1 import make_pdf
from datetime import datetime
from utils import setspace
from utils import calc_excel_hash, jsonmetainfo
from versioncontrol import ensure_version_entry,get_latest_version,is_latest_version_for_file_sheet
import sqlite3
import re


# COMMENT_TAGS = {"#", "//", "コメント", "comment"}
COMMENT_TAGS = {"コメント", "comment"}  # 完全一致で無視する語

qpattern = None

def get_nenji_by_subno(sub_no):
    """subNoからnenjiを取得"""
    db_path = "/Volumes/NBPlan/TTC/カルテ管理/2025/DB/classdb.db"
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT nennji FROM class WHERE subNo = ?", (sub_no,))
        result = cursor.fetchone()
        return str(result[0]) if result else None
    except sqlite3.Error as e:
        print(f"エラー: {e}")
        return None
    finally:
        conn.close()


def is_comment(tag: str) -> bool:
    """行タグがコメント行かどうかを判定"""
    if not tag:
        return True  # 空行もスキップ
    tag = tag.strip()
    if tag =="#ansbreak":
        return False
    if tag.startswith("#") or tag.startswith("//"):
        return True
    if tag in COMMENT_TAGS:
        return True
    return False

def get_cbn(num):
    if 1 <= num <= 26:
        return chr(ord('A') + num - 1)
    else:
        return None

def n2char(num):
    if isinstance(num, str):
        print(num)
        clst=num.split(",")
        dlst=[]
        for w in clst:
            print(w)
            w=int(w)
            # w=sdict[int(w)]
#            print(w)
            dlst.append(get_cbn(w))
        dlst.sort()
        return ','.join(dlst)
    else:
        return get_cbn(num)
        e=get_cbn(sdict[v])  

#-----------------------------------------------------------------------
def build_answer_columns(excel_path, nenji,sh,shname, version="A"):
    global qpattern

    # wb = load_workbook(excel_path)
    # sh = wb[shname]

    #hash値の作成
    ehash=calc_excel_hash(sh)
    print(ehash)

    # すでにハッシュが存在するかチェックし、なければ新規登録。
    wver=ensure_version_entry(ehash,str(excel_path),shname)
    
    # 解答jsonのhashとversionDBのhashが同じバージョンかチェックする
    cjinfo = jsonmetainfo(curdir / "work",f"answers_{shname}.json")  #jsonのhashを取得

    if is_latest_version_for_file_sheet(cjinfo[0], str(excel_path),shname) and 1!=1:
        # エクセルの最新バージョンと解答jsonが同じならばjsonは生成しない
        return None

    else:
        dt = datetime.now()
        metainfo={"type":"metainfo", "hash":ehash, "createdatetime":dt.strftime('%Y-%m-%d %H:%M:%S'), "verno":wver , "inputpath":str(excel_path), "sheetname":shname}

        # ===== ヘッダ情報 =====
        title = None        # examtitle の2列目
        subject = None      # subject の2列目
        fsyear = None       # subject の3列目
        kaito_message = None  # ansnote の2列目
        ans_width = None    # anssize の2列目（幅）
        ans_height = None   # anssize の3列目（高さ）

        # ===== 文脈（現在の問題など）=====
        qnum = 0
        current_koumoku = None    # b_question の2列目
        in_subgroup = False       # b_subgroup〜e_subgroup 内か
        sub_index = 0             # b_subquest ごとに1から振る

        # ===== 列詰め =====
        columns = []              # 出力する列（辞書）の配列
        def new_column():
            return {"width": [], "label": [], "answer": [], "height": [], "point": [], "koumoku": []}

        # まずは線形の回答リストを作る（Excelの出現順）
        answers_linear = []

        # --- B版並び替え用（b_question D列=orderB） ---
        q_orderB = {}            # {元qnum: orderB}
        current_orderB = None    # 現在の問題の orderB
        seq = 0                  # answers_linear の安定順序

        def _to_int_or_none(v):
            if v is None:
                return None
            s = str(v).strip()
            if s == "":
                return None
            try:
                return int(s)
            except Exception:
                try:
                    return int(float(s))
                except Exception:
                    return None

        def _parse_versions_cell(v):
            """セルが 'A' / 'B' / 'A,B' のとき set を返す。空なら None（全適用）"""
            if v is None:
                return None
            s = str(v).strip().upper()
            if not s:
                return None
            parts = re.split(r"[,\s]+", s)
            return {p for p in parts if p}
        selflg=False
        for row in sh.iter_rows(min_row=1, max_row=sh.max_row, values_only=True):

            tag = (row[0] or "").strip() if row[0] is not None else ""
    #        if tag in COMMENT_TAGS or tag == "":
    #            continue
            if is_comment(tag):
                continue

            # --- 強制改行用の制御タグ（#ansbreak） ---
            if tag == "#ansbreak":
                # B列に 'A' / 'B' / 'A,B' を書いた場合、そのバージョンだけ有効
                allowed = _parse_versions_cell(row[1] if len(row) > 1 else None)
                if allowed is not None and version.upper() not in allowed:
                    continue

                seq += 1
                answers_linear.append({
                    "label": "%%ansbreak%%",
                    "answer": "",
                    "koumoku": "",
                    "point": 0,
                    "width": 99.9,
                    "height": 1.0,
                    "_qnum": qnum,    # 直前の問題に紐づける（B版で並べ替えても一緒に移動）
                    "_seq": seq,
                })
                continue  # #ansbreak自体の処理は不要

            if tag == "examtitle":
                title = row[1] if len(row) > 1 else title

            elif tag == "subject":
                subject = str(row[1]) if len(row) > 1 and row[1] is not None else subject

            elif tag == "fsyear":
                fsyear = str(row[1]) if len(row) > 1 and row[1] is not None else fsyear

            elif tag == "ansnote":
                kaito_message = row[1] if len(row) > 1 else kaito_message

            elif tag == "anssize":
                # anssize 2列目=(幅,高さ)
                ansline_width, ansline_height = list(setspace(row[1], "ANSSIZE"))

            elif tag == "qpattern":
                qpattern = [x.strip() for x in row[1].split(",") if x.strip()]
                # versionmode の判定
                if not qpattern or qpattern == ["A"]:
                    versionmode = "single"
                else:
                    versionmode = "multi"

            elif tag == "b_question":
                qnum += 1
                current_koumoku = row[1] if len(row) > 1 else None
                in_subgroup = False
                sub_index = 0

                # B版は b_question のD列(orderB)で並べ替える（問題用紙と同じ仕様）
                if version != "A":
                    ob = _to_int_or_none(row[3] if len(row) > 3 else None)  # D列
                    if ob is None:
                        raise ValueError(f"b_question のD列(orderB)が空です: 元番号={qnum}")
                    current_orderB = ob
                else:
                    current_orderB = qnum

                q_orderB[qnum] = current_orderB

            elif tag == "b_subgroup":
                in_subgroup = True
                sub_index = 0

            elif tag == "b_subquest":
                sub_index += 1  # 小問番号は1から

            elif tag == "e_subgroup":
                in_subgroup = False
                sub_index = 0

            elif version != "A" and tag in ("b_select", "b_subselect"):
                select_order = {}
                sidx=0

            elif version != "A" and tag in ("select", "subselect"):
                if len(row) > 6 and row[6]:
                    try:
                        order_num = int(row[6])
                    except:
                        order_num = None
                sidx+=1
                select_order[sidx]=order_num
#------------------------------------------------------------------------------
            # elif version != "A" and tag in ("e_select", "e_subselect"):
            #     selflg=True

            # elif tag in ("e_answer", "e_subanswer"):
            #     selflg=False
            elif tag in ("b_answer", "b_subanswer"):
                #print(row[1])
                if row[1] == "#select":  #選択用の解答？
                    selflg=True
                else:
                    selflg=False
#------------------------------------------------------------------------------

            elif tag in ("answer", "subanswer"):
                '''
                A列：0 tag
                B列：1 答え
                C列：2 得点
                D列：3 幅,高さ 倍数 (w,h）
                E列：4 ラベル（番号の後につける）
                F列：5 
                G列：6 ver2用の選択問題の問題番号（シャッフル）と対応した解答番号
                '''
                # … 通常の answer 登録処理 …
    #            ans_text = row[1] if len(row) > 1 else ""
    
                ans_text = row[1]
                if row[1] is None:
                    ans_text = ""
                else:
                    t = str(ans_text)
#------------------------------------------------------------------------------
                    if selflg:
                        if version != "A":
                            ans_text = row[6]  #シャッフルされた解答
                        t=n2char(ans_text)
                        #selflg=False
#------------------------------------------------------------------------------
                    if "\n" in t:
                        # 改行を含む場合は <br /> に変換
                        ans_text = t.replace(' ', '&nbsp;').replace('\n', '<br />')
                    else:
                        ans_text = t
                # 点数
                point = 0
                if len(row) > 2 and row[2] is not None:
                    try:
                        point = int(row[2])
                    except Exception:
                        try:
                            point = int(str(row[2]).strip())
                        except Exception:
                            point = 0
                # 解答欄幅/高
                ans_width, ah = setspace(row[3],"ANSWH")
                ans_height = ansline_height * ah
                # ラベル決定
                if tag == "subanswer":
                    # 小問 → qnum-sub_index（Excelのラベル列は無視）
                    sub_label = str(row[4]).strip() if len(row) > 4 and row[4] else "" #None
                    label = f"{qnum}-{sub_index}{sub_label}"
                else:
                    # 大問 → qnum または qnum-サブラベル
                    sub_label = str(row[4]).strip() if len(row) > 4 and row[4] else None
                    if sub_label:
                        label = f"{qnum}-{sub_label}"
                    else:
                        label = str(qnum)
                # ★ pointが3点以上なら末尾に※
                try:
                    if float(point) >= 3 and not label.endswith("*"):
                        label += "*"
                except Exception:
                    pass

                # 問題の項目
                koumoku_val = row[5] if len(row) > 5  and row[5] else current_koumoku #(current_koumoku or "")

                seq += 1
                answers_linear.append({
                    "label": label,
                    "answer": ans_text,
                    "koumoku": koumoku_val,
                    "point": point,
                    "width": ans_width,
                    "height": ans_height,

                    # B版の並べ替え＆再番号付けに必要な情報
                    "_qnum": qnum,
                    "_is_sub": (tag == "subanswer"),
                    "_sub_index": sub_index,
                    "_suffix": (str(row[4]).strip() if len(row) > 4 and row[4] else ""),
                    "_seq": seq,
                })

            # それ以外のタグは無視（PAGEBREAK などは解答用紙には影響なし）
        if ansline_width is None or ansline_height is None:
            raise ValueError("anssize が見つからないか、幅/高さが数値ではありません（anssize 2列目=幅, 3列目=高さ）")


        # -----------------------------
        # B版：orderBで並べ替え → 問番号を1..Nに振り直す（問題用紙と合わせる）
        # -----------------------------
        if version != "A" and q_orderB:
            # orderB重複チェック
            vals = list(q_orderB.values())
            if len(vals) != len(set(vals)):
                # 重複値を表示
                dup = sorted({x for x in vals if vals.count(x) > 1})
                raise ValueError(f"b_question のD列(orderB)が重複しています: {dup}")

            # 元qnum -> 新しい問番号（1..N）
            ordered = sorted(q_orderB.items(), key=lambda kv: kv[1])  # (元qnum, orderB)
            newnum = {orig_qnum: i + 1 for i, (orig_qnum, _) in enumerate(ordered)}

            # answers_linear を orderB順に並べ替え（同一問題内は _seq で維持）
            def _sort_key(ent):
                oq = ent.get("_qnum", 10**9)
                ob = q_orderB.get(oq, 10**9)
                return (ob, ent.get("_seq", 10**9))
            answers_linear.sort(key=_sort_key)

            # ラベルを再構成（B版の問番号に合わせる）
            for ent in answers_linear:
                if ent.get("label") == "%%ansbreak%%":
                    continue
                oq = ent.get("_qnum")
                nq = newnum.get(oq)
                if nq is None:
                    continue

                suffix = ent.get("_suffix", "")
                if ent.get("_is_sub"):
                    ent["label"] = f"{nq}-{ent.get('_sub_index', 0)}{suffix}"
                else:
                    ent["label"] = f"{nq}-{suffix}" if suffix else str(nq)

                # ★ pointが3点以上なら末尾に※（B版の最終ラベルに反映）
                try:
                    if float(ent.get("point", 0)) >= 3 and not ent["label"].endswith("*"):
                        ent["label"] += "*"
                except Exception:
                    pass

            # 付加情報を落とす（出力JSONを綺麗に）
            for ent in answers_linear:
                ent.pop("_qnum", None)
                ent.pop("_is_sub", None)
                ent.pop("_sub_index", None)
                ent.pop("_suffix", None)
                ent.pop("_seq", None)
        # 列の最大件数（旧ロジック互換）
        max_per_col = max(1, int(500 / ansline_width))

        # 列に詰める（上→下、満杯になったら次の列）
        cur_col = new_column()
        cur_count = 0
        for ent in answers_linear:
            if cur_count + ent["width"] > max_per_col:
                columns.append(cur_col)
                cur_col = new_column()
                cur_count = 0
            if ent["label"] != "%%ansbreak%%":
                cur_col["width"].append(ent["width"])      # マス
                cur_col["height"].append(ent["height"])    # セル高さ
                cur_col["label"].append(ent["label"])      # "1" / "1-1" など
                cur_col["answer"].append(ent["answer"])    # "A", "②" など
                cur_col["point"].append(ent["point"])      # 点数
                cur_col["koumoku"].append(ent["koumoku"])  # 科目内カテゴリ

                cur_count += ent["width"]

            if cur_count == max_per_col:
                columns.append(cur_col)
                cur_col = new_column()
                cur_count = 0
        # 余りがあれば最後に出す
        if any(v for v in cur_col.values()):
            columns.append(cur_col)

        # 出力：ヘッダ + 列配列
        out = []
        out.append({
            "title": title,
            "nenji" : nenji,
            "width": ansline_width,
            "height": ansline_height,
            "kaito_message": kaito_message,
            "subject": subject,
            "fsyear": fsyear,
            "qversion":"",
            "metainfo":metainfo
        })
        out.extend(columns)
        return out

if __name__ == "__main__":
    import argparse
    from pathlib import Path

    def parse_cli_versions(v: str):
        """--version A / B / A,B を {'A','B'} にする。未指定や空は None（全出力）"""
        if v is None:
            return None
        s = str(v).strip().upper()
        if not s:
            return None
        parts = re.split(r"[,\s]+", s)
        return {p for p in parts if p}

    parser = argparse.ArgumentParser(description="試験問題.xlsx → 解答用JSON → 解答用紙PDF")
    parser.add_argument("subject", help="科目番号（=Excelシート名）")
    parser.add_argument("--version", default=None, help="出力するバージョン（A / B / A,B）。未指定なら全バージョン")
    args = parser.parse_args()

    subject = args.subject
    requested_versions = parse_cli_versions(args.version)  # Noneなら全バージョン

    # 現在の場所
    curdir = Path(__file__).parent.parent

    # 試験用のエクセルファイル
    examdata = "試験問題.xlsx"
    pt = curdir / "input" / examdata
    po = curdir / "work" / f"answers_{subject}.json"

    wb = load_workbook(pt)
    sh = wb[subject]

    nenji = get_nenji_by_subno(subject)

    # まずAで1回作って qpattern/versionmode を確定（既存の流れ）
    data = build_answer_columns(pt, nenji, sh, subject)

    if data:
        # outjsonの初期化
        outjson = {"versionmode": "", "versions": []}

        # blockを生成（single/multi）
        if qpattern is None or qpattern == ["A"]:
            outjson["versionmode"] = "single"
            outjson["versions"].append({
                "version": "A",
                "questions": data
            })
        else:
            outjson["versionmode"] = "multi"
            for ver in qpattern:
                print(f"★ バージョン {ver} を生成します")
                data_ver = build_answer_columns(pt, nenji, sh, subject, ver)
                outjson["versions"].append({
                    "version": ver,
                    "questions": data_ver
                })

        # JSONを出力（全バージョン分を保持）
        with open(po, "w", encoding="utf-8") as f:
            json.dump(outjson, f, ensure_ascii=False, indent=2)

        print(f"\n ✅ 出力: {po}")
    else:
        with open(po, "r", encoding="utf-8") as f:
            outjson = json.load(f)
        print("\n⏭️ エクセルのバージョンと解答jsonのバージョンは同じです")
        print(f"    {str(po)}を読みこみました")

    # ---- ここからPDF出力（要求仕様） ----
    available_versions = [v["version"] for v in outjson.get("versions", [])]

    if requested_versions is None:
        versions_to_render = available_versions
    else:
        versions_to_render = [v for v in available_versions if v.upper() in requested_versions]

    if not versions_to_render:
        print("❌ 指定された --version が、利用可能なバージョンに存在しません。")
        print(f"    利用可能: {available_versions}")
        raise SystemExit(1)

    for v in outjson["versions"]:
        ver = v["version"]
        if ver not in versions_to_render:
            continue

        # 出力パス：output/{subject}/{version}/
        outdir = curdir / "output" / subject / ver
        outdir.mkdir(parents=True, exist_ok=True)

        # ファイル名：{subject}_{version}_解答用紙.pdf
        pdfout = outdir / f"{subject}_{ver}_解答用紙.pdf"

        # PDF生成
        if outjson["versionmode"] == "single":
            make_pdf(v["questions"], pdfout, 7)
        else:
            make_pdf(v["questions"], pdfout, 7, ver)

        print(f"✅ PDF出力: {pdfout}")

