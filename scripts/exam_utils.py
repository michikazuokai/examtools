#!/usr/bin/env python3
# /Volumes/NBPlan/TTC/examtools/scripts/exam_utils.py

from __future__ import annotations

from pathlib import Path
from dataclasses import dataclass
from typing import Any
import argparse
import hashlib
import json
import re
import sys

import openpyxl
import yaml


# ============================================================
# Common @TTC utils
# ============================================================
#
# このファイル:
#   /Volumes/NBPlan/TTC/examtools/scripts/exam_utils.py
#
# EXAMTOOLS_SCRIPT_DIR = /Volumes/NBPlan/TTC/examtools/scripts
# EXAMTOOLS_ROOT       = /Volumes/NBPlan/TTC/examtools
# TTC_ROOT             = /Volumes/NBPlan/TTC
# COMMON_UTIL_DIR      = /Volumes/NBPlan/TTC/@TTC/util

EXAMTOOLS_SCRIPT_DIR = Path(__file__).resolve().parent
EXAMTOOLS_ROOT = EXAMTOOLS_SCRIPT_DIR.parent
TTC_ROOT = EXAMTOOLS_ROOT.parent

COMMON_UTIL_DIR = TTC_ROOT / "@TTC" / "util"

if str(COMMON_UTIL_DIR) not in sys.path:
    sys.path.insert(0, str(COMMON_UTIL_DIR))

try:
    from utils import (
        get_current_fsyear as common_get_current_fsyear,
        load_slideinfo_by_subno as common_load_slideinfo_by_subno,
        save_slideinfo as common_save_slideinfo,
    )
except ImportError as e:
    print("❌ 共通 utils.py の読み込みに失敗しました。", file=sys.stderr)
    print(f"COMMON_UTIL_DIR: {COMMON_UTIL_DIR}", file=sys.stderr)
    print(f"error: {e}", file=sys.stderr)
    raise


# ============================================================
# Dataclass
# ============================================================

@dataclass
class ExamContext:
    subject: str
    fsyear: str
    excel_path: Path
    work_dir: Path
    exam_koma_no: str
    sub_folder: Path

    # 追加
    exam_dir: Path

    # 互換・別名
    year: str = ""
    subject_dir: Path | None = None
    sheetname: str = ""

    wb: Any | None = None
    workbook: Any | None = None
    ws: Any | None = None
    worksheet: Any | None = None

    qpattern: str | None = None
    excel_hash: str | None = None

# ============================================================
# Argument helpers
# ============================================================

def add_subject_arg(parser: argparse.ArgumentParser) -> None:
    """
    各スクリプト共通の科目番号引数を追加する。
    """
    parser.add_argument(
        "subject",
        help="科目番号。例: 1020701",
    )


def add_dryrun_arg(parser: argparse.ArgumentParser) -> None:
    """
    dry-run 引数を追加する。
    旧コード互換のため、属性名は args.dryrun にする。
    """
    parser.add_argument(
        "--dry-run",
        "--dryrun",
        dest="dryrun",
        action="store_true",
        help="実際の書き込みを行わず、確認のみ実行する",
    )

# ============================================================
# Common delegation
# ============================================================

def get_current_fsyear() -> str:
    """
    現在年度を取得する。

    実体は /Volumes/NBPlan/TTC/@TTC/util/utils.py に委譲する。
    """
    return common_get_current_fsyear()


def load_slideinfo_by_subno(
    target_sub_no: str,
    target_year: str | None = None,
) -> tuple[dict[str, Any], Path]:
    """
    科目番号・年度から slideinfo.yaml を読み込む。

    実体は /Volumes/NBPlan/TTC/@TTC/util/utils.py に委譲する。
    """
    year = target_year or get_current_fsyear()
    return common_load_slideinfo_by_subno(str(target_sub_no), str(year))


def save_slideinfo(
    sub_folder: str | Path,
    slideinfo_data: dict[str, Any],
) -> None:
    """
    科目フォルダ配下の slideinfo/slideinfo.yaml を保存する。

    実体は /Volumes/NBPlan/TTC/@TTC/util/utils.py に委譲する。
    """
    common_save_slideinfo(Path(sub_folder), slideinfo_data)


# ============================================================
# Exam path helpers
# ============================================================

def get_exam_path(
    target_sub_no: str,
    target_year: str | None = None,
) -> tuple[Path, Path, str, Path]:
    """
    科目番号から、試験問題.xlsx と work フォルダを取得する。

    slideinfo.yaml の中で

        schedule_type: 試験

    になっているコマを試験回として探す。

    戻り値:
        excel_path, work_dir, exam_koma_no, subject_dir
    """
    year = target_year or get_current_fsyear()
    slideinfo_data, subject_dir = load_slideinfo_by_subno(target_sub_no, year)

    exam_koma_no: str | None = None

    for koma_key, info in slideinfo_data.items():
        if isinstance(info, dict) and info.get("schedule_type") == "試験":
            exam_koma_no = str(koma_key).zfill(2)
            break

    if exam_koma_no is None:
        raise ValueError(
            f"slideinfo.yaml 内に schedule_type: 試験 のコマが見つかりません。"
            f" subject={target_sub_no}, year={year}"
        )

    exam_dir = subject_dir / exam_koma_no
    excel_path = exam_dir / "試験問題.xlsx"
    work_dir = exam_dir / "work"

    return excel_path, work_dir, exam_koma_no, subject_dir

def load_exam_context(
    subject: str,
    *,
    load_workbook: bool = False,
    data_only: bool = False,
    sheetname: str | None = None,
) -> ExamContext:
    """
    科目番号から試験問題作成用の共通コンテキストを作る。
    旧 utils.py 互換の属性名もセットする。
    """
    fsyear = get_current_fsyear()
    excel_path, work_dir, exam_koma_no, sub_folder = get_exam_path(subject, fsyear)

    wb = None
    ws = None

    resolved_sheetname = sheetname or str(subject)

    if load_workbook:
        if not excel_path.exists():
            raise FileNotFoundError(f"試験問題.xlsx が見つかりません: {excel_path}")

        wb = openpyxl.load_workbook(excel_path, data_only=data_only)

        if resolved_sheetname in wb.sheetnames:
            ws = wb[resolved_sheetname]
        else:
            raise KeyError(
                f"Excel内にシート '{resolved_sheetname}' が見つかりません。"
                f" 使用可能なシート: {wb.sheetnames}"
            )

    qpattern = None
    excel_hash = None

    if excel_path.exists():
        try:
            qpattern = get_qpattern(excel_path)
        except Exception:
            qpattern = None

        try:
            excel_hash = calc_excel_hash(excel_path)
        except Exception:
            excel_hash = None
    exam_dir = Path(sub_folder) / str(exam_koma_no)

    return ExamContext(
        subject=str(subject),
        fsyear=str(fsyear),
        year=str(fsyear),

        excel_path=Path(excel_path),
        work_dir=Path(work_dir),
        exam_koma_no=str(exam_koma_no),

        sub_folder=Path(sub_folder),
        subject_dir=Path(sub_folder),

        # 追加
        exam_dir=Path(exam_dir),

        sheetname=resolved_sheetname,

        wb=wb,
        workbook=wb,
        ws=ws,
        worksheet=ws,

        qpattern=qpattern,
        excel_hash=excel_hash,
    )

def ensure_work_dir(work_dir: str | Path) -> Path:
    """
    work フォルダを作成して返す。
    """
    work_dir = Path(work_dir)
    work_dir.mkdir(parents=True, exist_ok=True)
    return work_dir


# ============================================================
# slideinfo.yaml update helpers for examtools
# ============================================================

def write_exam_path_to_slideinfo(
    subject: str,
    target_year: str | None,
    key_name: str,
    file_path: str | Path,
) -> Path:
    """
    科目別 slideinfo.yaml の exam -> exam -> key_name にパスを書き込む。

    例:
        exam:
          exam:
            exam_json: /path/to/work/1020701_exam.json
            ans_json: /path/to/work/1020701_ans.json
            exam_pdf: /path/to/16/1020701_A.pdf

    戻り値:
        保存した slideinfo.yaml のパス
    """
    year = target_year or get_current_fsyear()
    slideinfo_data, subject_dir = load_slideinfo_by_subno(subject, year)

    if "exam" not in slideinfo_data or slideinfo_data["exam"] is None:
        slideinfo_data["exam"] = {}

    if not isinstance(slideinfo_data["exam"], dict):
        raise ValueError("slideinfo.yaml の exam キーが辞書ではありません。")

    if "exam" not in slideinfo_data["exam"] or slideinfo_data["exam"]["exam"] is None:
        slideinfo_data["exam"]["exam"] = {}

    if not isinstance(slideinfo_data["exam"]["exam"], dict):
        raise ValueError("slideinfo.yaml の exam -> exam が辞書ではありません。")

    slideinfo_data["exam"]["exam"][str(key_name)] = str(Path(file_path))

    save_slideinfo(subject_dir, slideinfo_data)

    return Path(subject_dir) / "slideinfo" / "slideinfo.yaml"


# ============================================================
# Subject metadata helpers
# ============================================================

def get_nenji_by_subno(sub_no: str, target_year: str) -> str | None:
    """
    subNoから受講年次を取得する。
    旧版互換：slideinfo.yaml の target_year を読む。
    """
    try:
        slideinfo, _ = load_slideinfo_by_subno(sub_no, target_year)

        nenji = slideinfo.get("target_year")

        if nenji is None:
            print(f"slideinfo.yaml に target_year がありません: sub_no={sub_no}")
            return None

        return str(nenji)

    except Exception as e:
        print(f"年次情報の取得エラー: {e}")
        return None


# ============================================================
# Excel / hash helpers
# ============================================================

def calc_excel_hash(sheet) -> str:
    """
    Worksheet の内容からハッシュ値を計算する。
    旧 examtools/scripts/utils.py 互換。

    validate_excel.py / make_json.py / make_anspdf.py は、
    Excelファイルのパスではなく Worksheet を渡している。
    """
    content = []

    for row in sheet.iter_rows(values_only=True):
        content.append(",".join("" if v is None else str(v) for v in row))

    return hashlib.md5("\n".join(content).encode("utf-8")).hexdigest()

def get_qpattern(sheet) -> str:
    """
    Worksheet から qpattern を取得する。
    旧 examtools/scripts/utils.py 互換。

    想定：
      - A列に 'qpattern' があり、B列に値がある
      - または、どこかのセルに 'qpattern' があり、右隣のセルに値がある
      - 見つからない場合は 'A'
    """
    # 1. A列='qpattern'、B列=値 を優先
    for row in sheet.iter_rows():
        key_cell = row[0] if len(row) >= 1 else None
        val_cell = row[1] if len(row) >= 2 else None

        key = "" if key_cell is None or key_cell.value is None else str(key_cell.value).strip().lower()

        if key == "qpattern":
            if val_cell is not None and val_cell.value is not None and str(val_cell.value).strip():
                return str(val_cell.value).strip()

    # 2. 念のため、任意セル='qpattern'、右隣=値 も見る
    for row in sheet.iter_rows():
        for i, cell in enumerate(row):
            key = "" if cell.value is None else str(cell.value).strip().lower()

            if key == "qpattern":
                if i + 1 < len(row):
                    right_cell = row[i + 1]
                    if right_cell.value is not None and str(right_cell.value).strip():
                        return str(right_cell.value).strip()

    # 3. 見つからない場合
    return "A"


# ============================================================
# JSON metadata helpers
# ============================================================

def jsonmetainfo(
    *,
    subject: str,
    year: str,
    excel_path: str | Path,
    qpattern: str | None = None,
    extra: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """
    JSON出力用のメタ情報を作る。
    make_json.py などから利用する想定。
    """
    excel_path = Path(excel_path)

    meta = {
        "subject": str(subject),
        "year": str(year),
        "excel_path": str(excel_path),
        "excel_hash": calc_excel_hash(excel_path) if excel_path.exists() else "",
        "qpattern": qpattern or get_qpattern(excel_path),
    }

    if extra:
        meta.update(extra)

    return meta


# ============================================================
# Text parsing helpers
# ============================================================

def setspace(text, p1):
    defaults = {"ANSSIZE": (50.0, 60.0), "SPACEB_A": (0.0, 0.0), "ANSWH": (1.0, 1.0)}
    DEFAULTb, DEFAULTa = defaults.get(p1, (0.0, 0.0))

    def parse_num(val, default):
        try:
            return float(val) if val.strip() else default
        except ValueError:
            return default

    # 入力が文字列でなければデフォルト
    if not isinstance(text, str):
        return DEFAULTb, DEFAULTa

    # ( ) で囲まれていなければデフォルト
    text = text.strip()
    if not (text.startswith("(") and text.endswith(")")):
        return DEFAULTb, DEFAULTa

    # () 中身を分割
    parts = text[1:-1].split(",")

    # --- ANSWH専用ルール ---
    if p1 == "ANSWH":
        if len(parts) == 1:   # 例: (3)
            return parse_num(parts[0], 1.0), 1.0
        elif len(parts) == 2: # 例: (,3)
            return parse_num(parts[0], 1.0), parse_num(parts[1], 1.0)
        return 1.0, 1.0

    # --- 共通ルール ---
    if len(parts) != 2:
        return DEFAULTb, DEFAULTa
    return parse_num(parts[0], DEFAULTb), parse_num(parts[1], DEFAULTa)

def parse_with_number(value, default=None):
    """
    文字列から数値を取り出す。
    旧 examtools/scripts/utils.py 互換。

    使い方:
        wimg = parse_with_number(params[0], 0.85)

    例:
        "0.8"   -> 0.8
        "4"     -> 4
        "[0.8]" -> 0.8
        "S"     -> default
        ""      -> default
        None    -> default
    """
    if value is None:
        return default

    text = str(value).strip()

    if text == "":
        return default

    # [0.85] のような形式にも対応
    m = re.search(r"\[([0-9]+(?:\.[0-9]+)?)\]", text)
    if m:
        num_text = m.group(1)
    else:
        num_text = text

    try:
        if "." in num_text:
            return float(num_text)
        return int(num_text)
    except ValueError:
        return default


def parse_number_in_brackets(value: Any) -> tuple[str, int | float | None]:
    """
    parse_with_number の別名。
    今後こちらの名前に寄せたい場合の互換用。
    """
    return parse_with_number(value)


# ============================================================
# File helpers
# ============================================================

def load_json(path: str | Path) -> Any:
    """
    JSONファイルを読み込む。
    """
    path = Path(path)

    if not path.exists():
        raise FileNotFoundError(f"JSONファイルが見つかりません: {path}")

    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def save_json(path: str | Path, data: Any) -> None:
    """
    JSONファイルを保存する。
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def load_yaml(path: str | Path) -> dict[str, Any]:
    """
    examtools内で必要な場合の簡易YAMLロード。
    共通 utils.py と同名だが、examtools内の補助として残す。
    """
    path = Path(path)

    if not path.exists():
        raise FileNotFoundError(f"YAMLファイルが見つかりません: {path}")

    with open(path, "r", encoding="utf-8") as f:
        data = yaml.safe_load(f) or {}

    if not isinstance(data, dict):
        raise TypeError(f"YAMLの内容がdictではありません: {path}")

    return data


def save_yaml(path: str | Path, data: dict[str, Any]) -> None:
    """
    examtools内で必要な場合の簡易YAML保存。
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    with open(path, "w", encoding="utf-8") as f:
        yaml.safe_dump(data, f, allow_unicode=True, sort_keys=False)


# ============================================================
# Simple test
# ============================================================

if __name__ == "__main__":
    print("=" * 70)
    print("exam_utils.py 動作確認")
    print("=" * 70)

    try:
        print("\n[1] パス確認")
        print("EXAMTOOLS_SCRIPT_DIR:", EXAMTOOLS_SCRIPT_DIR)
        print("EXAMTOOLS_ROOT      :", EXAMTOOLS_ROOT)
        print("TTC_ROOT            :", TTC_ROOT)
        print("COMMON_UTIL_DIR     :", COMMON_UTIL_DIR)
        print("COMMON_UTIL exists  :", COMMON_UTIL_DIR.exists())

        print("\n[2] 年度確認")
        year = get_current_fsyear()
        print("fsyear:", year)

        # 必要に応じて変更してください。
        subject = "1020701"

        print("\n[3] 科目 slideinfo 読み込み確認")
        slideinfo_data, subject_dir = load_slideinfo_by_subno(subject, year)
        print("subject      :", subject)
        print("subject_dir  :", subject_dir)
        print("slideinfo keys sample:", list(slideinfo_data.keys())[:10])

        print("\n[4] 試験パス確認")
        try:
            excel_path, work_dir, exam_koma_no, sub_folder = get_exam_path(subject, year)
            print("exam_koma_no:", exam_koma_no)
            print("excel_path  :", excel_path)
            print("excel exists:", excel_path.exists())
            print("work_dir    :", work_dir)
            print("sub_folder  :", sub_folder)
        except Exception as e:
            print("試験パス確認はスキップされました:", e)

        print("\n✅ exam_utils.py の確認が完了しました。")

    except Exception as e:
        print("\n❌ 動作確認中にエラーが発生しました。")
        print(type(e).__name__, ":", e)
        raise

# ============================================================
# Simple test
# ============================================================

if __name__ == "__main__":
    print("=" * 70)
    print("exam_utils.py 単体テスト")
    print("=" * 70)

    try:
        # ------------------------------------------------------------
        # 1. パス確認
        # ------------------------------------------------------------
        print("\n[1] パス確認")
        print("EXAMTOOLS_SCRIPT_DIR:", EXAMTOOLS_SCRIPT_DIR)
        print("EXAMTOOLS_ROOT      :", EXAMTOOLS_ROOT)
        print("TTC_ROOT            :", TTC_ROOT)
        print("COMMON_UTIL_DIR     :", COMMON_UTIL_DIR)
        print("COMMON_UTIL exists  :", COMMON_UTIL_DIR.exists())

        if not COMMON_UTIL_DIR.exists():
            raise FileNotFoundError(f"共通utilフォルダが見つかりません: {COMMON_UTIL_DIR}")

        # ------------------------------------------------------------
        # 2. 年度取得
        # ------------------------------------------------------------
        print("\n[2] 年度取得")
        fsyear = get_current_fsyear()
        print("fsyear:", fsyear)

        if not fsyear:
            raise ValueError("fsyear が空です。")

        # ------------------------------------------------------------
        # 3. テスト対象
        # ------------------------------------------------------------
        # 実際に存在する科目番号に変更してください。
        subject = "1020701"

        print("\n[3] テスト対象")
        print("subject:", subject)

        # ------------------------------------------------------------
        # 4. slideinfo.yaml 読み込み
        # ------------------------------------------------------------
        print("\n[4] slideinfo.yaml 読み込み")
        slideinfo_data, subject_dir = load_slideinfo_by_subno(subject, fsyear)

        print("subject_dir:", subject_dir)
        print("exists     :", subject_dir.exists())
        print("type       :", type(slideinfo_data).__name__)
        print("keys sample:", list(slideinfo_data.keys())[:10])

        if not subject_dir.exists():
            raise FileNotFoundError(f"科目フォルダが見つかりません: {subject_dir}")

        if not isinstance(slideinfo_data, dict):
            raise TypeError("slideinfo_data が dict ではありません。")

        # ------------------------------------------------------------
        # 5. 試験回フォルダの取得
        # ------------------------------------------------------------
        print("\n[5] get_exam_path() 確認")

        excel_path, work_dir, exam_koma_no, sub_folder = get_exam_path(subject, fsyear)

        print("exam_koma_no:", exam_koma_no)
        print("excel_path  :", excel_path)
        print("excel exists:", excel_path.exists())
        print("work_dir    :", work_dir)
        print("work exists :", work_dir.exists())
        print("sub_folder  :", sub_folder)

        if not exam_koma_no:
            raise ValueError("exam_koma_no が取得できません。")

        if not excel_path.exists():
            print("⚠️ 試験問題.xlsx が見つかりません。")
            print("   ただし、schedule_type: 試験 のコマ検出までは成功しています。")

        # ------------------------------------------------------------
        # 6. load_exam_context() 確認
        # ------------------------------------------------------------
        print("\n[6] load_exam_context() 確認")

        ctx = load_exam_context(subject)

        print("ctx.subject     :", ctx.subject)
        print("ctx.year        :", ctx.year)
        print("ctx.excel_path  :", ctx.excel_path)
        print("ctx.work_dir    :", ctx.work_dir)
        print("ctx.exam_koma_no:", ctx.exam_koma_no)
        print("ctx.subject_dir :", ctx.subject_dir)

        if ctx.subject != str(subject):
            raise ValueError("ctx.subject が想定と異なります。")

        if ctx.year != str(fsyear):
            raise ValueError("ctx.year が想定と異なります。")

        # ------------------------------------------------------------
        # 7. qpattern 確認
        # ------------------------------------------------------------
        print("\n[7] get_qpattern() 確認")

        if ctx.excel_path.exists():
            qpattern = get_qpattern(ctx.excel_path)
            print("qpattern:", qpattern)
        else:
            print("⚠️ Excelが存在しないため get_qpattern() はスキップしました。")

        # ------------------------------------------------------------
        # 8. Excel hash 確認
        # ------------------------------------------------------------
        print("\n[8] calc_excel_hash() 確認")

        if ctx.excel_path.exists():
            excel_hash = calc_excel_hash(ctx.excel_path)
            print("excel_hash:", excel_hash)
            print("hash length:", len(excel_hash))

            if len(excel_hash) != 64:
                raise ValueError("SHA256ハッシュの長さが64ではありません。")
        else:
            print("⚠️ Excelが存在しないため calc_excel_hash() はスキップしました。")

        # ------------------------------------------------------------
        # 9. parse_with_number() 確認
        # ------------------------------------------------------------
        print("\n[9] parse_with_number() 確認")

        samples = [
            "image.png[4.5]",
            "sample.tex[10]",
            "normal text",
            "",
            None,
        ]

        for s in samples:
            parsed_text, number = parse_with_number(s)
            print(f"{s!r} -> text={parsed_text!r}, number={number!r}")

        # ------------------------------------------------------------
        # 10. setspace() 確認
        # ------------------------------------------------------------
        print("\n[10] setspace() 確認")

        print("setspace(None) :", repr(setspace(None)))
        print("setspace('abc'):", repr(setspace("abc")))
        print("setspace(123)  :", repr(setspace(123)))

        # ------------------------------------------------------------
        # 11. jsonmetainfo() 確認
        # ------------------------------------------------------------
        print("\n[11] jsonmetainfo() 確認")

        if ctx.excel_path.exists():
            meta = jsonmetainfo(
                subject=ctx.subject,
                year=ctx.year,
                excel_path=ctx.excel_path,
            )
            print("meta keys:", list(meta.keys()))
            print("subject  :", meta.get("subject"))
            print("year     :", meta.get("year"))
            print("qpattern :", meta.get("qpattern"))
        else:
            print("⚠️ Excelが存在しないため jsonmetainfo() はスキップしました。")

        # ------------------------------------------------------------
        # 12. get_nenji_by_subno() 確認
        # ------------------------------------------------------------
        print("\n[12] get_nenji_by_subno() 確認")

        nenji = get_nenji_by_subno(subject, fsyear)
        print("nenji:", nenji)

        # ------------------------------------------------------------
        # 13. 保存系テスト
        # ------------------------------------------------------------
        print("\n[13] 保存系テスト")
        print("write_exam_path_to_slideinfo() は slideinfo.yaml を書き換えるため、")
        print("この単体テストでは実行していません。")

        # 実際に保存テストしたい場合だけ、下のコメントを外してください。
        #
        # test_path = ctx.work_dir / "_test_dummy.json"
        # saved_slideinfo_path = write_exam_path_to_slideinfo(
        #     subject=ctx.subject,
        #     target_year=ctx.year,
        #     key_name="_test_dummy",
        #     file_path=test_path,
        # )
        # print("saved_slideinfo_path:", saved_slideinfo_path)

        print("\n" + "=" * 70)
        print("✅ exam_utils.py 単体テスト完了")
        print("=" * 70)

    except Exception as e:
        print("\n" + "=" * 70)
        print("❌ exam_utils.py 単体テスト失敗")
        print("=" * 70)
        print(type(e).__name__, ":", e)
        raise